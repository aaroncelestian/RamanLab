#!/usr/bin/env python3
"""
Mineral Modes Database Browser Qt6 Version - Enhanced
Qt6 conversion of the mineral database browser for RamanLab
Enhanced with dielectric tensor viewing, improved plotting, and better UX
"""

import os
import pickle
import numpy as np
import sys
from pathlib import Path
import pandas as pd

# Qt6 imports
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QTabWidget,
    QLabel, QPushButton, QLineEdit, QTextEdit, QComboBox, QSpinBox, QDoubleSpinBox,
    QGroupBox, QSplitter, QFileDialog, QMessageBox, QProgressBar,
    QStatusBar, QMenuBar, QTableWidget, QTableWidgetItem, QHeaderView,
    QDialog, QFormLayout, QDialogButtonBox, QListWidget, QListWidgetItem,
    QScrollArea, QFrame, QCheckBox, QTreeWidget, QTreeWidgetItem, QGridLayout
)
from PySide6.QtCore import Qt, QTimer, QStandardPaths
from PySide6.QtGui import QAction, QFont, QPixmap, QColor

# Import matplotlib for Qt6
import matplotlib
matplotlib.use("QtAgg")
import matplotlib.pyplot as plt
from matplotlib.figure import Figure

try:
    from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
    from core.matplotlib_config import CompactNavigationToolbar as NavigationToolbar
except ImportError:
    # Fallback - use PySide6/Qt6 compatible backend
    from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.backends.backend_qtagg import NavigationToolbar2QT as NavigationToolbar

from core.matplotlib_config import configure_compact_ui, apply_theme


class MineralModesDatabaseQt6(QMainWindow):
    """Qt6 Mineral Modes Database Browser - Enhanced Version."""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        
        # Apply compact UI configuration for consistent toolbar sizing
        apply_theme('compact')
        
        self.setWindowTitle("RamanLab - Calculated Raman Character Database Browser")
        self.setMinimumSize(1200, 800)
        self.resize(1400, 900)
        
        # Initialize database
        self.database_path = os.path.join(os.path.dirname(__file__), "mineral_modes.pkl")
        self.database = {}
        self.current_mineral = None
        
        # Set up UI first (including status bar)
        self.setup_ui()
        self.setup_menu_bar()
        self.setup_status_bar()
        
        # Load database after UI is set up
        self.load_database()
        
        # Update mineral list
        self.update_mineral_list()
    
    def setup_ui(self):
        """Set up the main user interface."""
        # Create central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Create main layout (horizontal splitter)
        main_layout = QHBoxLayout(central_widget)
        main_splitter = QSplitter(Qt.Horizontal)
        main_layout.addWidget(main_splitter)
        
        # Left panel - mineral list and controls
        self.setup_left_panel(main_splitter)
        
        # Right panel - modes display and visualization
        self.setup_right_panel(main_splitter)
        
        # Set splitter proportions (30% left, 70% right)
        main_splitter.setSizes([420, 980])
    
    def setup_left_panel(self, parent):
        """Set up the left panel with mineral list and controls."""
        left_frame = QFrame()
        left_layout = QVBoxLayout(left_frame)
        
        # Search group
        search_group = QGroupBox("Search Minerals")
        search_layout = QVBoxLayout(search_group)
        
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Search mineral names...")
        self.search_edit.textChanged.connect(self.filter_mineral_list)
        search_layout.addWidget(self.search_edit)
        
        clear_search_btn = QPushButton("Clear Search")
        clear_search_btn.clicked.connect(self.clear_search)
        search_layout.addWidget(clear_search_btn)
        
        left_layout.addWidget(search_group)
        
        # Filter controls
        filter_group = QGroupBox("Filter by Properties")
        filter_layout = QVBoxLayout(filter_group)
        
        # Crystal system filter
        crystal_filter_layout = QHBoxLayout()
        crystal_filter_layout.addWidget(QLabel("Crystal System:"))
        self.crystal_filter_combo = QComboBox()
        self.crystal_filter_combo.addItems([
            "All", "Cubic", "Tetragonal", "Orthorhombic", 
            "Hexagonal", "Trigonal", "Monoclinic", "Triclinic"
        ])
        self.crystal_filter_combo.currentTextChanged.connect(self.filter_mineral_list)
        crystal_filter_layout.addWidget(self.crystal_filter_combo)
        filter_layout.addLayout(crystal_filter_layout)
        
        # Has dielectric data filter
        self.has_dielectric_check = QCheckBox("Has Dielectric Tensor Data")
        self.has_dielectric_check.toggled.connect(self.filter_mineral_list)
        filter_layout.addWidget(self.has_dielectric_check)
        
        # Minimum modes filter
        min_modes_layout = QHBoxLayout()
        min_modes_layout.addWidget(QLabel("Min Raman Modes:"))
        self.min_modes_spin = QSpinBox()
        self.min_modes_spin.setRange(0, 100)
        self.min_modes_spin.setValue(0)
        self.min_modes_spin.valueChanged.connect(self.filter_mineral_list)
        min_modes_layout.addWidget(self.min_modes_spin)
        filter_layout.addLayout(min_modes_layout)
        
        left_layout.addWidget(filter_group)
        
        # Mineral list group
        list_group = QGroupBox("Mineral List")
        list_layout = QVBoxLayout(list_group)
        
        self.mineral_list = QListWidget()
        self.mineral_list.currentItemChanged.connect(self.on_mineral_selected)
        list_layout.addWidget(self.mineral_list)
        
        # Add mineral count label
        self.mineral_count_label = QLabel("Total: 0 minerals")
        list_layout.addWidget(self.mineral_count_label)
        
        left_layout.addWidget(list_group)
        
        # Controls group
        controls_group = QGroupBox("Database Operations")
        controls_layout = QVBoxLayout(controls_group)
        
        # Add mineral button
        add_btn = QPushButton("Add New Mineral")
        add_btn.clicked.connect(self.add_mineral)
        controls_layout.addWidget(add_btn)
        
        # Edit mineral button
        edit_btn = QPushButton("Edit Selected Mineral")
        edit_btn.clicked.connect(self.edit_mineral)
        controls_layout.addWidget(edit_btn)
        
        # Delete mineral button
        delete_btn = QPushButton("Delete Selected Mineral")
        delete_btn.clicked.connect(self.delete_mineral)
        delete_btn.setStyleSheet("QPushButton { background-color: #f57c00; color: white; }")
        controls_layout.addWidget(delete_btn)
        
        # Save database button
        save_btn = QPushButton("Save Database")
        save_btn.clicked.connect(self.save_database)
        save_btn.setStyleSheet("QPushButton { background-color: #1976d2; color: white; }")
        controls_layout.addWidget(save_btn)
        
        # Database statistics
        stats_group = QGroupBox(" ") #removed Database Statistics name to make it look cleaner
        stats_layout = QVBoxLayout(stats_group)
        
        self.stats_text = QTextEdit()
        self.stats_text.setMaximumHeight(120)
        self.stats_text.setReadOnly(True)
        stats_layout.addWidget(self.stats_text)
        
        refresh_stats_btn = QPushButton("Refresh Statistics")
        refresh_stats_btn.clicked.connect(self.update_database_stats)
        stats_layout.addWidget(refresh_stats_btn)
        
        controls_layout.addWidget(stats_group)
        left_layout.addWidget(controls_group)
        left_layout.addStretch()
        
        parent.addWidget(left_frame)
    
    def setup_right_panel(self, parent):
        """Set up the right panel with modes display and visualization."""
        right_frame = QFrame()
        right_layout = QVBoxLayout(right_frame)
        
        # Mineral information header
        self.setup_mineral_info_header(right_layout)
        
        # Create tab widget for different views
        self.tab_widget = QTabWidget()
        
        # Calculated spectrum tab (first tab - main view)
        self.spectrum_tab = self.create_spectrum_tab()
        self.tab_widget.addTab(self.spectrum_tab, "Calculated Spectrum")
        
        # Modes table tab
        self.modes_tab = self.create_modes_tab()
        self.tab_widget.addTab(self.modes_tab, "Raman Modes Table")
        
        # Dielectric tensor tab
        self.dielectric_tab = self.create_dielectric_tab()
        self.tab_widget.addTab(self.dielectric_tab, "Dielectric Tensors")
        
        # Advanced properties tab
        self.advanced_tab = self.create_advanced_tab()
        self.tab_widget.addTab(self.advanced_tab, "Advanced Properties")
        
        right_layout.addWidget(self.tab_widget)
        parent.addWidget(right_frame)
    
    def setup_mineral_info_header(self, parent_layout):
        """Set up the mineral information header similar to database browser."""
        info_group = QGroupBox("Selected Mineral Information")
        info_layout = QGridLayout(info_group)
        
        # Create info labels
        row = 0
        info_layout.addWidget(QLabel("Name:"), row, 0)
        self.name_label = QLabel("")
        self.name_label.setStyleSheet("border: 1px solid gray; padding: 2px; background: white; font-weight: bold;")
        info_layout.addWidget(self.name_label, row, 1)
        
        info_layout.addWidget(QLabel("Crystal System:"), row, 2)
        self.crystal_system_label = QLabel("")
        self.crystal_system_label.setStyleSheet("border: 1px solid gray; padding: 2px; background: white;")
        info_layout.addWidget(self.crystal_system_label, row, 3)
        
        row += 1
        info_layout.addWidget(QLabel("Total Modes:"), row, 0)
        self.total_modes_label = QLabel("")
        self.total_modes_label.setStyleSheet("border: 1px solid gray; padding: 2px; background: white;")
        info_layout.addWidget(self.total_modes_label, row, 1)
        
        info_layout.addWidget(QLabel("Raman Active:"), row, 2)
        self.raman_active_label = QLabel("")
        self.raman_active_label.setStyleSheet("border: 1px solid gray; padding: 2px; background: white;")
        info_layout.addWidget(self.raman_active_label, row, 3)
        
        # Control buttons
        button_layout = QVBoxLayout()
        view_detailed_btn = QPushButton("View Detailed Info")
        view_detailed_btn.clicked.connect(self.view_detailed_mineral_info)
        button_layout.addWidget(view_detailed_btn)
        
        export_spectrum_btn = QPushButton("Export Spectrum")
        export_spectrum_btn.clicked.connect(self.export_calculated_spectrum)
        button_layout.addWidget(export_spectrum_btn)
        
        info_layout.addLayout(button_layout, 0, 4, 2, 1)
        
        parent_layout.addWidget(info_group)
    
    def create_spectrum_tab(self):
        """Create the calculated spectrum visualization tab (main tab)."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Visualization controls
        controls_group = QGroupBox("Spectrum Visualization Controls")
        controls_layout = QHBoxLayout(controls_group)
        
        # Line width control
        controls_layout.addWidget(QLabel("Line Width:"))
        self.line_width_spin = QDoubleSpinBox()
        self.line_width_spin.setRange(0.5, 5.0)
        self.line_width_spin.setSingleStep(0.5)
        self.line_width_spin.setValue(2.0)
        self.line_width_spin.valueChanged.connect(self.update_plot)
        controls_layout.addWidget(self.line_width_spin)
        
        # Peak width control
        controls_layout.addWidget(QLabel("Peak Width:"))
        self.peak_width_spin = QDoubleSpinBox()
        self.peak_width_spin.setRange(1.0, 50.0)
        self.peak_width_spin.setSingleStep(1.0)
        self.peak_width_spin.setValue(10.0)
        self.peak_width_spin.valueChanged.connect(self.update_plot)
        controls_layout.addWidget(self.peak_width_spin)
        
        # Intensity threshold for labeling
        controls_layout.addWidget(QLabel("Label Threshold (%):"))
        self.label_threshold_spin = QDoubleSpinBox()
        self.label_threshold_spin.setRange(0.0, 50.0)
        self.label_threshold_spin.setSingleStep(1.0)
        self.label_threshold_spin.setValue(5.0)  # Default 5% threshold
        self.label_threshold_spin.valueChanged.connect(self.update_plot)
        controls_layout.addWidget(self.label_threshold_spin)
        
        # Show labels checkbox
        self.show_labels_check = QCheckBox("Show Peak Labels")
        self.show_labels_check.setChecked(True)
        self.show_labels_check.toggled.connect(self.update_plot)
        controls_layout.addWidget(self.show_labels_check)
        
        # Filter Raman-active only checkbox
        self.raman_only_check = QCheckBox("Raman Active Only")
        self.raman_only_check.setChecked(True)
        self.raman_only_check.toggled.connect(self.update_plot)
        controls_layout.addWidget(self.raman_only_check)
        
        controls_layout.addStretch()
        
        layout.addWidget(controls_group)
        
        # Create matplotlib figure and canvas
        self.figure = Figure(figsize=(12, 8))
        self.canvas = FigureCanvas(self.figure)
        self.toolbar = NavigationToolbar(self.canvas, tab)
        
        # Create the plot
        self.ax = self.figure.add_subplot(111)
        self.ax.set_xlabel("Wavenumber (cm⁻¹)")
        self.ax.set_ylabel("Normalized Intensity")
        self.ax.set_title("Calculated Raman Spectrum")
        self.ax.grid(True, alpha=0.3)
        
        layout.addWidget(self.toolbar)
        layout.addWidget(self.canvas)
        
        return tab
    
    def create_modes_tab(self):
        """Create the Raman modes table tab."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Modes group
        modes_group = QGroupBox("Raman Modes")
        modes_layout = QVBoxLayout(modes_group)
        
        # Create modes table
        self.modes_table = QTableWidget()
        self.modes_table.setColumnCount(4)
        self.modes_table.setHorizontalHeaderLabels([
            "Position (cm⁻¹)", "Symmetry", "Intensity", "Raman Active"
        ])
        self.modes_table.horizontalHeader().setStretchLastSection(True)
        self.modes_table.setAlternatingRowColors(True)
        modes_layout.addWidget(self.modes_table)
        
        # Mode controls
        mode_controls = QHBoxLayout()
        
        add_mode_btn = QPushButton("Add Mode")
        add_mode_btn.clicked.connect(self.add_mode)
        mode_controls.addWidget(add_mode_btn)
        
        edit_mode_btn = QPushButton("Edit Mode")
        edit_mode_btn.clicked.connect(self.edit_mode)
        mode_controls.addWidget(edit_mode_btn)
        
        delete_mode_btn = QPushButton("Delete Mode")
        delete_mode_btn.clicked.connect(self.delete_mode)
        delete_mode_btn.setStyleSheet("QPushButton { background-color: #d32f2f; color: white; }")
        mode_controls.addWidget(delete_mode_btn)
        
        mode_controls.addStretch()
        modes_layout.addLayout(mode_controls)
        
        layout.addWidget(modes_group)
        
        return tab
    
    def create_dielectric_tab(self):
        """Create the dielectric tensor viewing tab."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Dielectric tensor group
        tensor_group = QGroupBox("Dielectric Tensors")
        tensor_layout = QVBoxLayout(tensor_group)
        
        # Controls for tensor selection
        controls_layout = QHBoxLayout()
        
        controls_layout.addWidget(QLabel("Tensor Type:"))
        self.tensor_type_combo = QComboBox()
        self.tensor_type_combo.addItems([
            "High-frequency (ε∞)", "Static (ε₀)", "Born Effective Charges", "Refractive Index (n)"
        ])
        self.tensor_type_combo.currentTextChanged.connect(self.update_dielectric_display)
        controls_layout.addWidget(self.tensor_type_combo)
        
        controls_layout.addStretch()
        
        view_full_btn = QPushButton("View Full Tensor Data")
        view_full_btn.clicked.connect(self.view_full_dielectric_data)
        controls_layout.addWidget(view_full_btn)
        
        tensor_layout.addLayout(controls_layout)
        
        # Tensor display table
        self.dielectric_table = QTableWidget()
        self.dielectric_table.setColumnCount(4)
        self.dielectric_table.setHorizontalHeaderLabels(["Component", "X", "Y", "Z"])
        self.dielectric_table.setMaximumHeight(200)
        tensor_layout.addWidget(self.dielectric_table)
        
        # Tensor properties text
        self.tensor_properties_text = QTextEdit()
        self.tensor_properties_text.setReadOnly(True)
        self.tensor_properties_text.setMaximumHeight(150)
        tensor_layout.addWidget(self.tensor_properties_text)
        
        layout.addWidget(tensor_group)
        
        # Optical properties group
        optical_group = QGroupBox("Calculated Optical Properties")
        optical_layout = QVBoxLayout(optical_group)
        
        self.optical_properties_text = QTextEdit()
        self.optical_properties_text.setReadOnly(True)
        optical_layout.addWidget(self.optical_properties_text)
        
        layout.addWidget(optical_group)
        
        return tab
    
    def create_advanced_tab(self):
        """Create the advanced properties tab."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Phonon modes analysis
        phonon_group = QGroupBox("Phonon Mode Analysis")
        phonon_layout = QVBoxLayout(phonon_group)
        
        self.phonon_analysis_text = QTextEdit()
        self.phonon_analysis_text.setReadOnly(True)
        phonon_layout.addWidget(self.phonon_analysis_text)
        
        layout.addWidget(phonon_group)
        
        # Crystal structure information
        structure_group = QGroupBox("Crystal Structure")
        structure_layout = QVBoxLayout(structure_group)
        
        self.structure_info_text = QTextEdit()
        self.structure_info_text.setReadOnly(True)
        structure_layout.addWidget(self.structure_info_text)
        
        layout.addWidget(structure_group)
        
        return tab
    
    def setup_menu_bar(self):
        """Set up the menu bar."""
        menubar = self.menuBar()
        
        # File menu
        file_menu = menubar.addMenu("File")
        
        load_action = QAction("Load Database", self)
        load_action.triggered.connect(self.load_database_file)
        file_menu.addAction(load_action)
        
        save_action = QAction("Save Database", self)
        save_action.triggered.connect(self.save_database)
        file_menu.addAction(save_action)
        
        file_menu.addSeparator()
        
        export_action = QAction("Export Database", self)
        export_action.triggered.connect(self.export_database)
        file_menu.addAction(export_action)
        
        import_action = QAction("Import Database", self)
        import_action.triggered.connect(self.import_database)
        file_menu.addAction(import_action)
        
        file_menu.addSeparator()
        
        exit_action = QAction("Exit", self)
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        
        # Tools menu
        tools_menu = menubar.addMenu("Tools")
        
        inspect_action = QAction("Inspect Database Structure", self)
        inspect_action.triggered.connect(self.inspect_database_structure)
        tools_menu.addAction(inspect_action)
        
        tools_menu.addSeparator()
        
        compare_action = QAction("Compare Minerals", self)
        compare_action.triggered.connect(self.launch_mineral_comparison)
        tools_menu.addAction(compare_action)
        
        batch_analysis_action = QAction("Batch Analysis", self)
        batch_analysis_action.triggered.connect(self.launch_batch_analysis)
        tools_menu.addAction(batch_analysis_action)
        
        # Help menu
        help_menu = menubar.addMenu("Help")
        
        about_action = QAction("About", self)
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)
    
    def setup_status_bar(self):
        """Set up the status bar."""
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("Calculated Raman Character Database Browser - Ready")
    
    def load_database(self):
        """Load the mineral modes database."""
        try:
            if os.path.exists(self.database_path):
                with open(self.database_path, 'rb') as f:
                    self.database = pickle.load(f)
                if hasattr(self, 'status_bar'):
                    self.status_bar.showMessage(f"Loaded {len(self.database)} minerals from database")
            else:
                self.database = {}
                if hasattr(self, 'status_bar'):
                    self.status_bar.showMessage("Created new database (file not found)")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load database:\n{str(e)}")
            self.database = {}
        
        # Update statistics after loading
        self.update_database_stats()
    
    def save_database(self):
        """Save the mineral modes database."""
        try:
            with open(self.database_path, 'wb') as f:
                pickle.dump(self.database, f)
            if hasattr(self, 'status_bar'):
                self.status_bar.showMessage("Database saved successfully")
            QMessageBox.information(self, "Success", "Database saved successfully!")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save database:\n{str(e)}")
    
    def update_mineral_list(self):
        """Update the mineral list widget."""
        self.filter_mineral_list()
    
    def filter_mineral_list(self):
        """Filter the mineral list based on search text and filters."""
        search_text = self.search_edit.text().lower()
        crystal_filter = self.crystal_filter_combo.currentText()
        has_dielectric_filter = self.has_dielectric_check.isChecked()
        min_modes = self.min_modes_spin.value()
        
        # Clear and repopulate the list
        self.mineral_list.clear()
        filtered_minerals = []
        
        for mineral_name, mineral_data in self.database.items():
            if mineral_name.startswith('__'):
                continue
            
            # Skip non-dictionary values (metadata keys)
            if not isinstance(mineral_data, dict):
                continue
                
            # Apply search filter
            if search_text and search_text not in mineral_name.lower():
                continue
                
            # Apply crystal system filter
            crystal_system = mineral_data.get('crystal_system', '')
            if crystal_filter != "All" and crystal_system != crystal_filter:
                continue
                
            # Apply dielectric data filter
            if has_dielectric_filter and not self.has_dielectric_data(mineral_data):
                continue
                
            # Apply minimum modes filter - use new method
            modes = self.get_modes_for_mineral(mineral_name)
            if len(modes) < min_modes:
                continue
                
            # Add to filtered list if passes all filters
            filtered_minerals.append(mineral_name)
        
        # Sort alphabetically and add to widget
        filtered_minerals.sort()
        for mineral_name in filtered_minerals:
            self.mineral_list.addItem(mineral_name)
        
        # Update count
        self.mineral_count_label.setText(f"Showing: {len(filtered_minerals)} / {len(self.database)} minerals")
    
    def clear_search(self):
        """Clear the search field and reset filters."""
        self.search_edit.clear()
        self.crystal_filter_combo.setCurrentText("All")
        self.has_dielectric_check.setChecked(False)
        self.min_modes_spin.setValue(0)
        self.filter_mineral_list()
    
    def has_dielectric_data(self, mineral_data):
        """Check if mineral has dielectric tensor data."""
        # Look in various potential locations
        tensor_keys = [
            'dielectric_tensors', 'dielectric_tensor', 'eps_inf', 'eps_0',
            'born_charges', 'born_effective_charges', 'z_star', 'Z*'
        ]
        
        for key in tensor_keys:
            if key in mineral_data and mineral_data[key] is not None:
                return True
        
        return False
    
    def on_mineral_selected(self, current, previous):
        """Handle mineral selection change."""
        if current:
            mineral_name = current.text()
            self.current_mineral = mineral_name
            self.update_mineral_info_header(mineral_name)
            self.update_modes_table(mineral_name)
            self.update_dielectric_display()
            self.update_advanced_display()
            self.update_plot()
    
    def update_mineral_info_header(self, mineral_name):
        """Update the mineral information header."""
        if mineral_name not in self.database:
            return
        
        mineral_data = self.database[mineral_name]
        
        # Update labels
        self.name_label.setText(mineral_name)
        self.crystal_system_label.setText(mineral_data.get('crystal_system', 'Unknown'))
        
        # Use the new method that handles phonon_modes conversion
        modes = self.get_modes_for_mineral(mineral_name)
        self.total_modes_label.setText(str(len(modes)))
        
        # Count Raman active modes (exclude modes with 'u' in symmetry)
        raman_active = 0
        for mode in modes:
            if len(mode) >= 2:
                symmetry = str(mode[1]).lower()
                if 'u' not in symmetry:  # Raman active if no 'u'
                    raman_active += 1
        
        self.raman_active_label.setText(str(raman_active))
    
    def is_raman_active(self, symmetry):
        """Check if a mode is Raman active (no 'u' in symmetry)."""
        return 'u' not in str(symmetry).lower()
    
    def update_modes_table(self, mineral_name):
        """Update the modes table for the selected mineral."""
        self.modes_table.setRowCount(0)
        
        if mineral_name not in self.database:
            return
        
        # Use the new method that handles phonon_modes conversion
        modes = self.get_modes_for_mineral(mineral_name)
        
        self.modes_table.setRowCount(len(modes))
        
        for i, mode in enumerate(modes):
            if len(mode) >= 3:
                position, symmetry, intensity = mode[:3]
                
                # Position
                position_item = QTableWidgetItem(f"{position:.1f}")
                self.modes_table.setItem(i, 0, position_item)
                
                # Symmetry
                symmetry_item = QTableWidgetItem(str(symmetry))
                self.modes_table.setItem(i, 1, symmetry_item)
                
                # Intensity
                intensity_item = QTableWidgetItem(f"{intensity:.3f}")
                self.modes_table.setItem(i, 2, intensity_item)
                
                # Raman Active
                is_active = self.is_raman_active(symmetry)
                active_item = QTableWidgetItem("Yes" if is_active else "No")
                if not is_active:
                    active_item.setBackground(QColor(255, 200, 200))  # Light red background
                self.modes_table.setItem(i, 3, active_item)
    
    def update_plot(self):
        """Update the calculated Raman spectrum plot with enhanced peak labeling."""
        self.ax.clear()
        
        if not self.current_mineral or self.current_mineral not in self.database:
            self.ax.text(0.5, 0.5, 'Select a mineral to view calculated spectrum',
                        ha='center', va='center', transform=self.ax.transAxes,
                        fontsize=14, bbox=dict(boxstyle='round', facecolor='lightblue', alpha=0.8))
            self.canvas.draw()
            return
        
        # Use the new method that handles phonon_modes conversion
        modes = self.get_modes_for_mineral(self.current_mineral)
        
        if not modes:
            self.ax.text(0.5, 0.5, 'No calculated modes available for this mineral',
                        ha='center', va='center', transform=self.ax.transAxes,
                        fontsize=12, bbox=dict(boxstyle='round', facecolor='lightyellow', alpha=0.8))
            self.canvas.draw()
            return
        
        # Extract mode data and filter for Raman active modes if requested
        positions = []
        intensities = []
        symmetries = []
        
        for mode in modes:
            if len(mode) >= 3:
                pos, sym, intensity = mode[:3]
                
                # Filter Raman-active only if checkbox is checked
                if self.raman_only_check.isChecked():
                    if not self.is_raman_active(sym):
                        continue
                
                positions.append(float(pos))
                intensities.append(float(intensity))
                symmetries.append(str(sym))
        
        if not positions:
            self.ax.text(0.5, 0.5, 'No Raman-active modes to display',
                        ha='center', va='center', transform=self.ax.transAxes,
                        fontsize=12, bbox=dict(boxstyle='round', facecolor='lightcoral', alpha=0.8))
            self.canvas.draw()
            return
        
        # Normalize intensities
        max_intensity = max(intensities) if intensities else 1.0
        normalized_intensities = [i / max_intensity for i in intensities]
        
        # Create spectrum with Gaussian peaks
        x_min = min(positions) - 200
        x_max = max(positions) + 200
        x = np.linspace(x_min, x_max, 2000)
        
        spectrum = np.zeros_like(x)
        peak_width = self.peak_width_spin.value()
        
        for pos, norm_intensity in zip(positions, normalized_intensities):
            # Add Gaussian peak
            peak = norm_intensity * np.exp(-0.5 * ((x - pos) / peak_width) ** 2)
            spectrum += peak
        
        # Plot spectrum
        line_width = self.line_width_spin.value()
        self.ax.plot(x, spectrum, 'b-', linewidth=line_width, label='Calculated Raman Spectrum', alpha=0.8)
        
        # Plot individual peaks and labels
        label_threshold = self.label_threshold_spin.value() / 100.0  # Convert percentage to fraction
        
        for pos, norm_intensity, sym in zip(positions, normalized_intensities, symmetries):
            # Plot peak line
            color = 'red' if self.is_raman_active(sym) else 'gray'
            alpha = 1.0 if self.is_raman_active(sym) else 0.5
            
            self.ax.plot([pos, pos], [0, norm_intensity], color=color, 
                        linewidth=2, alpha=alpha)
            
            # Add labels if requested and above threshold
            if (self.show_labels_check.isChecked() and 
                norm_intensity >= label_threshold and 
                self.is_raman_active(sym)):  # Only label Raman-active modes
                
                self.ax.annotate(f'{pos:.0f}\n{sym}', 
                               xy=(pos, norm_intensity), 
                               xytext=(0, 15),
                               textcoords='offset points',
                               ha='center', va='bottom',
                               fontsize=9, fontweight='bold',
                               bbox=dict(boxstyle='round,pad=0.3', 
                                       facecolor='yellow', alpha=0.8, edgecolor='black'),
                               arrowprops=dict(arrowstyle='->', lw=1))
        
        self.ax.set_xlabel("Wavenumber (cm⁻¹)", fontsize=12)
        self.ax.set_ylabel("Normalized Intensity", fontsize=12)
        self.ax.set_title(f"Calculated Raman Spectrum: {self.current_mineral}", fontsize=14, fontweight='bold')
        self.ax.grid(True, alpha=0.3)
        self.ax.legend(fontsize=10)
        
        # Set reasonable y-limits
        self.ax.set_ylim(-0.05, 1.1)
        
        self.canvas.draw()
    
    def add_mineral(self):
        """Add a new mineral to the database."""
        dialog = MineralEditDialog(self)
        if dialog.exec() == QDialog.Accepted:
            mineral_data = dialog.get_mineral_data()
            name = mineral_data['name']
            
            if name in self.database:
                QMessageBox.warning(self, "Warning", f"Mineral '{name}' already exists!")
                return
            
            self.database[name] = mineral_data
            self.update_mineral_list()
            if hasattr(self, 'status_bar'):
                self.status_bar.showMessage(f"Added mineral: {name}")
    
    def edit_mineral(self):
        """Edit the selected mineral."""
        if not self.current_mineral:
            QMessageBox.warning(self, "Warning", "Please select a mineral to edit.")
            return
        
        mineral_data = self.database[self.current_mineral].copy()
        dialog = MineralEditDialog(self, mineral_data)
        if dialog.exec() == QDialog.Accepted:
            updated_data = dialog.get_mineral_data()
            self.database[self.current_mineral] = updated_data
            self.update_modes_table(self.current_mineral)
            self.update_info_display(self.current_mineral)
            self.update_plot()
            if hasattr(self, 'status_bar'):
                self.status_bar.showMessage(f"Updated mineral: {self.current_mineral}")
    
    def delete_mineral(self):
        """Delete the selected mineral."""
        if not self.current_mineral:
            QMessageBox.warning(self, "Warning", "Please select a mineral to delete.")
            return
        
        reply = QMessageBox.question(
            self, 
            "Confirm Delete", 
            f"Are you sure you want to delete '{self.current_mineral}'?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            del self.database[self.current_mineral]
            self.update_mineral_list()
            self.current_mineral = None
            self.modes_table.setRowCount(0)
            self.info_text.clear()
            self.update_plot()
            if hasattr(self, 'status_bar'):
                self.status_bar.showMessage("Mineral deleted")
    
    def add_mode(self):
        """Add a mode to the selected mineral."""
        if not self.current_mineral:
            QMessageBox.warning(self, "Warning", "Please select a mineral first.")
            return
        
        dialog = ModeEditDialog(self)
        if dialog.exec() == QDialog.Accepted:
            mode_data = dialog.get_mode_data()
            
            if 'modes' not in self.database[self.current_mineral]:
                self.database[self.current_mineral]['modes'] = []
            
            self.database[self.current_mineral]['modes'].append(mode_data)
            self.update_modes_table(self.current_mineral)
            self.update_info_display(self.current_mineral)
            self.update_plot()
            if hasattr(self, 'status_bar'):
                self.status_bar.showMessage("Mode added")
    
    def edit_mode(self):
        """Edit the selected mode."""
        current_row = self.modes_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "Warning", "Please select a mode to edit.")
            return
        
        if not self.current_mineral:
            return
        
        modes = self.database[self.current_mineral].get('modes', [])
        if current_row >= len(modes):
            return
        
        mode = modes[current_row]
        dialog = ModeEditDialog(self, mode)
        if dialog.exec() == QDialog.Accepted:
            updated_mode = dialog.get_mode_data()
            modes[current_row] = updated_mode
            self.update_modes_table(self.current_mineral)
            self.update_info_display(self.current_mineral)
            self.update_plot()
            if hasattr(self, 'status_bar'):
                self.status_bar.showMessage("Mode updated")
    
    def delete_mode(self):
        """Delete the selected mode."""
        current_row = self.modes_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "Warning", "Please select a mode to delete.")
            return
        
        if not self.current_mineral:
            return
        
        modes = self.database[self.current_mineral].get('modes', [])
        if current_row >= len(modes):
            return
        
        reply = QMessageBox.question(
            self, 
            "Confirm Delete", 
            "Are you sure you want to delete this mode?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            del modes[current_row]
            self.update_modes_table(self.current_mineral)
            self.update_info_display(self.current_mineral)
            self.update_plot()
            if hasattr(self, 'status_bar'):
                self.status_bar.showMessage("Mode deleted")
    
    def load_database_file(self):
        """Load a database from file."""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Load Mineral Database",
            QStandardPaths.writableLocation(QStandardPaths.DocumentsLocation),
            "Pickle files (*.pkl);;All files (*.*)"
        )
        
        if file_path:
            try:
                with open(file_path, 'rb') as f:
                    self.database = pickle.load(f)
                self.database_path = file_path
                self.update_mineral_list()
                if hasattr(self, 'status_bar'):
                    self.status_bar.showMessage(f"Loaded database from {Path(file_path).name}")
                QMessageBox.information(self, "Success", f"Loaded {len(self.database)} minerals")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to load database:\n{str(e)}")
    
    def export_database(self):
        """Export the database to a file with enhanced format support."""
        file_path, selected_filter = QFileDialog.getSaveFileName(
            self,
            "Export Mineral Database",
            QStandardPaths.writableLocation(QStandardPaths.DocumentsLocation),
            "Pickle files (*.pkl);;JSON files (*.json);;CSV files (*.csv);;SQLite Database (*.db);;Excel files (*.xlsx);;All files (*.*)"
        )
        
        if file_path:
            try:
                if file_path.endswith('.json'):
                    # Export as JSON
                    import json
                    # Handle pandas DataFrames in JSON export
                    def serialize_for_json(obj):
                        if isinstance(obj, pd.DataFrame):
                            return obj.to_dict('records')
                        elif hasattr(obj, 'tolist'):  # numpy arrays
                            return obj.tolist()
                        return str(obj)
                    
                    with open(file_path, 'w') as f:
                        json.dump(self.database, f, indent=2, default=serialize_for_json)
                
                elif file_path.endswith('.csv'):
                    # Export as CSV (flattened format)
                    import csv
                    with open(file_path, 'w', newline='') as f:
                        writer = csv.writer(f)
                        # Write header
                        writer.writerow(['Mineral_Name', 'Crystal_System', 'Point_Group', 'Space_Group', 
                                       'Mode_Position', 'Mode_Symmetry', 'Mode_Intensity', 'Raman_Active',
                                       'Has_Dielectric_Data', 'Total_Modes'])
                        
                        for mineral_name, mineral_data in self.database.items():
                            if mineral_name.startswith('__') or not isinstance(mineral_data, dict):
                                continue
                            
                            crystal_system = mineral_data.get('crystal_system', '')
                            point_group = mineral_data.get('point_group', '')
                            space_group = mineral_data.get('space_group', '')
                            has_dielectric = self.has_dielectric_data(mineral_data)
                            modes = mineral_data.get('modes', [])
                            
                            if modes:
                                for mode in modes:
                                    if len(mode) >= 3:
                                        pos, sym, intensity = mode[:3]
                                        raman_active = self.is_raman_active(sym)
                                        writer.writerow([mineral_name, crystal_system, point_group, space_group,
                                                       pos, sym, intensity, raman_active, has_dielectric, len(modes)])
                            else:
                                # Mineral with no modes
                                writer.writerow([mineral_name, crystal_system, point_group, space_group,
                                               '', '', '', '', has_dielectric, 0])
                
                elif file_path.endswith('.db'):
                    # Export as SQLite database
                    import sqlite3
                    conn = sqlite3.connect(file_path)
                    cursor = conn.cursor()
                    
                    # Create tables
                    cursor.execute('''
                        CREATE TABLE minerals (
                            id INTEGER PRIMARY KEY,
                            name TEXT UNIQUE,
                            crystal_system TEXT,
                            point_group TEXT,
                            space_group TEXT,
                            total_modes INTEGER,
                            has_dielectric_data BOOLEAN
                        )
                    ''')
                    
                    cursor.execute('''
                        CREATE TABLE modes (
                            id INTEGER PRIMARY KEY,
                            mineral_id INTEGER,
                            position REAL,
                            symmetry TEXT,
                            intensity REAL,
                            raman_active BOOLEAN,
                            FOREIGN KEY (mineral_id) REFERENCES minerals (id)
                        )
                    ''')
                    
                    cursor.execute('''
                        CREATE TABLE dielectric_tensors (
                            id INTEGER PRIMARY KEY,
                            mineral_id INTEGER,
                            tensor_type TEXT,
                            component TEXT,
                            x_value REAL,
                            y_value REAL,
                            z_value REAL,
                            FOREIGN KEY (mineral_id) REFERENCES minerals (id)
                        )
                    ''')
                    
                    # Insert data
                    for mineral_name, mineral_data in self.database.items():
                        if mineral_name.startswith('__') or not isinstance(mineral_data, dict):
                            continue
                        
                        has_dielectric = self.has_dielectric_data(mineral_data)
                        modes = mineral_data.get('modes', [])
                        
                        # Insert mineral
                        cursor.execute('''
                            INSERT INTO minerals (name, crystal_system, point_group, space_group, total_modes, has_dielectric_data)
                            VALUES (?, ?, ?, ?, ?, ?)
                        ''', (mineral_name, 
                              mineral_data.get('crystal_system', ''),
                              mineral_data.get('point_group', ''),
                              mineral_data.get('space_group', ''),
                              len(modes),
                              has_dielectric))
                        
                        mineral_id = cursor.lastrowid
                        
                        # Insert modes
                        for mode in modes:
                            if len(mode) >= 3:
                                pos, sym, intensity = mode[:3]
                                raman_active = self.is_raman_active(sym)
                                cursor.execute('''
                                    INSERT INTO modes (mineral_id, position, symmetry, intensity, raman_active)
                                    VALUES (?, ?, ?, ?, ?)
                                ''', (mineral_id, pos, sym, intensity, raman_active))
                        
                        # Insert dielectric tensor data if available
                        dielectric_data = None
                        if 'dielectric_tensors' in mineral_data:
                            data = mineral_data['dielectric_tensors']
                            if data is not None and not (isinstance(data, pd.DataFrame) and data.empty):
                                dielectric_data = data
                        
                        if dielectric_data is not None and isinstance(dielectric_data, pd.DataFrame):
                            for _, row in dielectric_data.iterrows():
                                cursor.execute('''
                                    INSERT INTO dielectric_tensors 
                                    (mineral_id, tensor_type, component, x_value, y_value, z_value)
                                    VALUES (?, ?, ?, ?, ?, ?)
                                ''', (mineral_id,
                                      row.get('Tensor', ''),
                                      row.get('Component', ''),
                                      float(row.get('X', 0)),
                                      float(row.get('Y', 0)),
                                      float(row.get('Z', 0))))
                    
                    conn.commit()
                    conn.close()
                
                elif file_path.endswith('.xlsx'):
                    # Export as Excel file
                    try:
                        import openpyxl
                    except ImportError:
                        QMessageBox.warning(self, "Missing Dependency", 
                                          "openpyxl is required for Excel export. Install with: pip install openpyxl")
                        return
                    
                    from openpyxl import Workbook
                    wb = Workbook()
                    
                    # Minerals sheet
                    ws_minerals = wb.active
                    ws_minerals.title = "Minerals"
                    ws_minerals.append(['Name', 'Crystal System', 'Point Group', 'Space Group', 
                                      'Total Modes', 'Raman Active Modes', 'Has Dielectric Data'])
                    
                    for mineral_name, mineral_data in self.database.items():
                        if mineral_name.startswith('__') or not isinstance(mineral_data, dict):
                            continue
                        
                        modes = mineral_data.get('modes', [])
                        raman_count = sum(1 for mode in modes if len(mode) >= 2 and self.is_raman_active(mode[1]))
                        
                        ws_minerals.append([
                            mineral_name,
                            mineral_data.get('crystal_system', ''),
                            mineral_data.get('point_group', ''),
                            mineral_data.get('space_group', ''),
                            len(modes),
                            raman_count,
                            self.has_dielectric_data(mineral_data)
                        ])
                    
                    # Modes sheet
                    ws_modes = wb.create_sheet("Modes")
                    ws_modes.append(['Mineral', 'Position', 'Symmetry', 'Intensity', 'Raman Active'])
                    
                    for mineral_name, mineral_data in self.database.items():
                        if mineral_name.startswith('__') or not isinstance(mineral_data, dict):
                            continue
                        
                        modes = mineral_data.get('modes', [])
                        for mode in modes:
                            if len(mode) >= 3:
                                pos, sym, intensity = mode[:3]
                                ws_modes.append([mineral_name, pos, sym, intensity, self.is_raman_active(sym)])
                    
                    wb.save(file_path)
                
                else:
                    # Export as pickle (default)
                    with open(file_path, 'wb') as f:
                        pickle.dump(self.database, f)
                
                QMessageBox.information(self, "Success", f"Database exported to {Path(file_path).name}")
                
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to export database:\n{str(e)}")
    
    def import_database(self):
        """Import and merge a database from file with enhanced format support."""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Import Mineral Database",
            QStandardPaths.writableLocation(QStandardPaths.DocumentsLocation),
            "Pickle files (*.pkl);;JSON files (*.json);;SQLite Database (*.db);;CSV files (*.csv);;Excel files (*.xlsx);;All files (*.*)"
        )
        
        if file_path:
            try:
                imported_data = {}
                
                if file_path.endswith('.json'):
                    # Import from JSON
                    import json
                    with open(file_path, 'r') as f:
                        imported_data = json.load(f)
                
                elif file_path.endswith('.db'):
                    # Import from SQLite database
                    import sqlite3
                    conn = sqlite3.connect(file_path)
                    cursor = conn.cursor()
                    
                    # Get all minerals
                    cursor.execute('SELECT * FROM minerals')
                    minerals = cursor.fetchall()
                    
                    for mineral_row in minerals:
                        mineral_id, name, crystal_system, point_group, space_group, total_modes, has_dielectric = mineral_row
                        
                        mineral_data = {
                            'crystal_system': crystal_system or '',
                            'point_group': point_group or '',
                            'space_group': space_group or '',
                            'modes': []
                        }
                        
                        # Get modes for this mineral
                        cursor.execute('SELECT position, symmetry, intensity, raman_active FROM modes WHERE mineral_id = ?', (mineral_id,))
                        modes = cursor.fetchall()
                        
                        for mode_row in modes:
                            position, symmetry, intensity, raman_active = mode_row
                            mineral_data['modes'].append((position, symmetry, intensity))
                        
                        # Get dielectric tensor data if available
                        cursor.execute('SELECT tensor_type, component, x_value, y_value, z_value FROM dielectric_tensors WHERE mineral_id = ?', (mineral_id,))
                        tensor_data = cursor.fetchall()
                        
                        if tensor_data:
                            # Convert to DataFrame format
                            import pandas as pd
                            df_data = []
                            for tensor_row in tensor_data:
                                tensor_type, component, x_val, y_val, z_val = tensor_row
                                df_data.append({
                                    'Tensor': tensor_type,
                                    'Component': component,
                                    'X': x_val,
                                    'Y': y_val,
                                    'Z': z_val
                                })
                            
                            if df_data:
                                mineral_data['dielectric_tensors'] = pd.DataFrame(df_data)
                        
                        imported_data[name] = mineral_data
                    
                    conn.close()
                
                elif file_path.endswith('.csv'):
                    # Import from CSV
                    import csv
                    with open(file_path, 'r') as f:
                        reader = csv.DictReader(f)
                        
                        for row in reader:
                            mineral_name = row.get('Mineral_Name', '')
                            if not mineral_name:
                                continue
                            
                            # Initialize mineral if not exists
                            if mineral_name not in imported_data:
                                imported_data[mineral_name] = {
                                    'crystal_system': row.get('Crystal_System', ''),
                                    'point_group': row.get('Point_Group', ''),
                                    'space_group': row.get('Space_Group', ''),
                                    'modes': []
                                }
                            
                            # Add mode if present
                            mode_pos = row.get('Mode_Position', '')
                            mode_sym = row.get('Mode_Symmetry', '')
                            mode_int = row.get('Mode_Intensity', '')
                            
                            if mode_pos and mode_sym and mode_int:
                                try:
                                    imported_data[mineral_name]['modes'].append((
                                        float(mode_pos), mode_sym, float(mode_int)
                                    ))
                                except ValueError:
                                    pass  # Skip invalid entries
                
                elif file_path.endswith('.xlsx'):
                    # Import from Excel
                    try:
                        import openpyxl
                    except ImportError:
                        QMessageBox.warning(self, "Missing Dependency", 
                                          "openpyxl is required for Excel import. Install with: pip install openpyxl")
                        return
                    
                    from openpyxl import load_workbook
                    wb = load_workbook(file_path)
                    
                    # Read minerals sheet
                    if 'Minerals' in wb.sheetnames:
                        ws_minerals = wb['Minerals']
                        # Skip header row
                        for row in ws_minerals.iter_rows(min_row=2, values_only=True):
                            if row[0]:  # If name is not empty
                                name = row[0]
                                imported_data[name] = {
                                    'crystal_system': row[1] or '',
                                    'point_group': row[2] or '',
                                    'space_group': row[3] or '',
                                    'modes': []
                                }
                    
                    # Read modes sheet
                    if 'Modes' in wb.sheetnames:
                        ws_modes = wb['Modes']
                        # Skip header row
                        for row in ws_modes.iter_rows(min_row=2, values_only=True):
                            if row[0] and row[1] and row[2] and row[3]:  # All required fields
                                mineral_name = row[0]
                                position = float(row[1])
                                symmetry = row[2]
                                intensity = float(row[3])
                                
                                if mineral_name in imported_data:
                                    imported_data[mineral_name]['modes'].append((position, symmetry, intensity))
                
                else:
                    # Import from pickle (default)
                    with open(file_path, 'rb') as f:
                        imported_data = pickle.load(f)
                
                # Merge with existing database
                conflicts = []
                for name, data in imported_data.items():
                    if name in self.database:
                        conflicts.append(name)
                
                if conflicts:
                    reply = QMessageBox.question(
                        self,
                        "Conflicts Found",
                        f"Found {len(conflicts)} conflicting minerals. Overwrite existing data?",
                        QMessageBox.Yes | QMessageBox.No,
                        QMessageBox.No
                    )
                    if reply == QMessageBox.No:
                        # Only import non-conflicting minerals
                        for conflict in conflicts:
                            if conflict in imported_data:
                                del imported_data[conflict]
                        
                        if not imported_data:
                            QMessageBox.information(self, "Import Cancelled", "No new minerals to import.")
                            return
                
                # Merge the data
                self.database.update(imported_data)
                self.update_mineral_list()
                self.update_database_stats()
                
                QMessageBox.information(
                    self, 
                    "Success", 
                    f"Imported {len(imported_data)} minerals from {Path(file_path).name}\n"
                    f"Total minerals in database: {len(self.database)}"
                )
                
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to import database:\n{str(e)}")
    
    def show_about(self):
        """Show about dialog."""
        about_text = """
        Mineral Modes Database Browser Qt6
        
        A tool for managing and visualizing mineral Raman modes.
        
        Features:
        • Browse and search mineral database
        • View and edit Raman modes
        • Visualize simulated spectra
        • Import/Export database files
        
        Part of RamanLab Qt6 Suite
        """
        QMessageBox.about(self, "About Mineral Modes Browser", about_text)
    
    def update_dielectric_display(self):
        """Update the dielectric tensor display."""
        self.dielectric_table.setRowCount(0)
        self.tensor_properties_text.clear()
        self.optical_properties_text.clear()
        
        if not self.current_mineral or self.current_mineral not in self.database:
            return
        
        mineral_data = self.database[self.current_mineral]
        tensor_type = self.tensor_type_combo.currentText()
        
        # Look for tensor data in various locations based on type
        dielectric_data = None
        
        if tensor_type.startswith("Refractive"):
            # Look specifically for refractive index data first
            if 'n_tensor' in mineral_data:
                n_tensor = mineral_data['n_tensor']
                if n_tensor is not None:
                    # Handle numpy array directly
                    self.display_n_tensor_data(n_tensor, tensor_type)
                    return
            
            # If no n_tensor, fall back to deriving from dielectric data
            for key in ['eps_inf', 'epsilon_inf', 'dielectric_tensors', 'dielectric_tensor']:
                if key in mineral_data:
                    data = mineral_data[key]
                    if data is not None and not (isinstance(data, pd.DataFrame) and data.empty):
                        dielectric_data = data
                        break
        
        elif tensor_type.startswith("Born"):
            # Look specifically for Born charges
            for key in ['born_charges', 'born_effective_charges', 'z_star', 'Z*', 'effective_charges']:
                if key in mineral_data:
                    data = mineral_data[key]
                    if data is not None and not (isinstance(data, pd.DataFrame) and data.empty):
                        dielectric_data = data
                        break
        
        # If no specific data found, look in general dielectric tensors
        if dielectric_data is None:
            if 'dielectric_tensors' in mineral_data:
                data = mineral_data['dielectric_tensors']
                if data is not None and not (isinstance(data, pd.DataFrame) and data.empty):
                    dielectric_data = data
            
            if dielectric_data is None and 'dielectric_tensor' in mineral_data:
                data = mineral_data['dielectric_tensor']
                if data is not None and not (isinstance(data, pd.DataFrame) and data.empty):
                    dielectric_data = data
        
        if dielectric_data is None:
            # Show what keys are available for debugging
            available_keys = [k for k in mineral_data.keys() if not k.startswith('_')]
            debug_msg = f"No dielectric tensor data available for this mineral.\n\n"
            debug_msg += f"Available data keys: {available_keys}\n"
            debug_msg += f"Looking for tensor type: {tensor_type}"
            self.tensor_properties_text.setPlainText(debug_msg)
            return
        
        # Debug: Show what we found
        debug_text = f"Found dielectric data type: {type(dielectric_data)}\n"
        if isinstance(dielectric_data, list) and len(dielectric_data) > 0:
            debug_text += f"List length: {len(dielectric_data)}\n"
            debug_text += f"Sample items (first 3):\n"
            for i, item in enumerate(dielectric_data[:3]):
                debug_text += f"  [{i}]: {item}\n"
        
        # Filter data based on tensor type selection
        filtered_data = self.filter_tensor_data_by_type(dielectric_data, tensor_type)
        
        if filtered_data is None:
            # Show debug info if available
            debug_message = debug_text + f"\n\nNo {tensor_type} data found in the available tensor data."
            if hasattr(self, '_debug_info'):
                debug_message += f"\n\n{self._debug_info}"
                delattr(self, '_debug_info')  # Clean up
            self.tensor_properties_text.setPlainText(debug_message)
            return
        
        # Parse and display tensor data
        if isinstance(filtered_data, pd.DataFrame):
            self.display_dataframe_tensors(filtered_data, tensor_type)
        elif isinstance(filtered_data, dict):
            self.display_dict_tensors(filtered_data, tensor_type)
        elif isinstance(filtered_data, list):
            self.display_list_tensors(filtered_data, tensor_type)
        else:
            self.tensor_properties_text.setPlainText(f"Tensor data format not supported: {type(filtered_data)}")
    
    def display_n_tensor_data(self, n_tensor, tensor_type):
        """Display refractive index tensor data from n_tensor numpy array."""
        try:
            import numpy as np
            
            # Convert to numpy array if not already
            if not isinstance(n_tensor, np.ndarray):
                n_tensor = np.array(n_tensor)
            
            # Ensure it's a 3x3 matrix
            if n_tensor.shape != (3, 3):
                self.tensor_properties_text.setPlainText(f"Invalid n_tensor shape: {n_tensor.shape}. Expected (3, 3).")
                return
            
            # Set up table for 3x3 matrix + eigenvalues
            self.dielectric_table.setRowCount(4)
            
            # Display the 3x3 refractive index matrix
            components = ['xx', 'yy', 'zz']
            for i in range(3):
                self.dielectric_table.setItem(i, 0, QTableWidgetItem(components[i]))
                for j in range(3):
                    value = float(n_tensor[i, j])
                    self.dielectric_table.setItem(i, j + 1, QTableWidgetItem(f"{value:.4f}"))
            
            # Calculate and display eigenvalues
            eigenvalues = np.linalg.eigvals(n_tensor)
            eigenvalues = np.sort(eigenvalues)[::-1]  # Sort descending
            
            self.dielectric_table.setItem(3, 0, QTableWidgetItem("Eigenvalues:"))
            for j, eig_val in enumerate(eigenvalues[:3]):
                self.dielectric_table.setItem(3, j + 1, QTableWidgetItem(f"{float(eig_val):.4f}"))
            
            # Calculate and display properties
            properties_text = f"Refractive Index Tensor Properties:\n"
            properties_text += "=" * 50 + "\n\n"
            
            properties_text += "Refractive Index Matrix:\n"
            for i, comp in enumerate(components):
                row_values = [f"{n_tensor[i, j]:.4f}" for j in range(3)]
                properties_text += f"  {comp}: [{row_values[0]}, {row_values[1]}, {row_values[2]}]\n"
            
            properties_text += f"\nPrincipal refractive indices (eigenvalues):\n"
            for i, eig_val in enumerate(eigenvalues):
                properties_text += f"  n{i+1}: {eig_val:.4f}\n"
            
            # Calculate optical properties
            valid_n = [n for n in eigenvalues if n > 0]
            if len(valid_n) >= 2:
                avg_n = np.mean(valid_n)
                max_n = max(valid_n)
                min_n = min(valid_n)
                birefringence = max_n - min_n
                
                properties_text += f"\nAverage refractive index: {avg_n:.4f}\n"
                properties_text += f"Birefringence (Δn): {birefringence:.4f}\n"
                
                # Determine optical character
                if len(valid_n) >= 3:
                    n1, n2, n3 = sorted(valid_n, reverse=True)
                    if abs(n1 - n2) < 0.001 and abs(n2 - n3) < 0.001:
                        optical_char = "Isotropic"
                    elif abs(n1 - n2) < 0.001:
                        optical_char = "Uniaxial negative" if n3 < n1 else "Uniaxial positive"
                    elif abs(n2 - n3) < 0.001:
                        optical_char = "Uniaxial positive" if n1 > n2 else "Uniaxial negative"
                    else:
                        optical_char = "Biaxial"
                        # Determine 2V angle for biaxial
                        beta = sorted(valid_n)[1]  # intermediate value
                        try:
                            if n1 != n3:
                                v_z = np.sqrt((n2**2 - n3**2) / (n1**2 - n3**2))
                                two_v = 2 * np.arcsin(v_z) * 180 / np.pi
                                optical_char += f" (2V ≈ {two_v:.1f}°)"
                        except:
                            pass
                    
                    properties_text += f"Optical character: {optical_char}\n"
            
            self.tensor_properties_text.setPlainText(properties_text)
            
            # Update optical properties display
            optical_text = f"Optical Properties from Refractive Index Tensor:\n\n"
            optical_text += f"Principal refractive indices:\n"
            for i, n in enumerate(eigenvalues):
                optical_text += f"  n{i+1}: {n:.4f}\n"
            
            if len(valid_n) >= 2:
                optical_text += f"\nAverage refractive index: {np.mean(valid_n):.4f}\n"
                optical_text += f"Birefringence: {max(valid_n) - min(valid_n):.4f}\n"
                
                # Add dielectric constant equivalents
                optical_text += f"\nEquivalent dielectric constants (ε = n²):\n"
                for i, n in enumerate(eigenvalues):
                    if n > 0:
                        eps = n**2
                        optical_text += f"  ε{i+1}: {eps:.4f}\n"
            
            self.optical_properties_text.setPlainText(optical_text)
            
        except ImportError:
            self.tensor_properties_text.setPlainText("NumPy required for refractive index calculations. Install with: pip install numpy")
        except Exception as e:
            self.tensor_properties_text.setPlainText(f"Error displaying refractive index tensor: {str(e)}")
    
    def filter_tensor_data_by_type(self, dielectric_data, tensor_type):
        """Filter tensor data based on the selected tensor type."""
        try:
            if isinstance(dielectric_data, list):
                # First, analyze what tensor types are available
                available_tensors = set()
                available_atoms = set()
                
                for item in dielectric_data:
                    if isinstance(item, dict):
                        tensor_key = item.get('Tensor', item.get('tensor', ''))
                        atom_key = item.get('Atom', item.get('atom', ''))
                        if tensor_key:
                            available_tensors.add(str(tensor_key))
                        if atom_key:
                            available_atoms.add(str(atom_key))
                
                # Debug info
                debug_info = f"Available tensor types: {sorted(available_tensors)}\n"
                debug_info += f"Available atoms: {sorted(available_atoms)}\n"
                debug_info += f"Total items: {len(dielectric_data)}\n\n"
                
                # Filter list based on tensor type
                filtered_items = []
                
                for item in dielectric_data:
                    if isinstance(item, dict):
                        # Check for tensor type indicators
                        tensor_key = item.get('Tensor', item.get('tensor', ''))
                        atom_key = item.get('Atom', item.get('atom', ''))
                        
                        if tensor_type.startswith("Born"):
                            # Look for Born charges - should have Atom field or Born-related tensor names
                            if (atom_key or 
                                'born' in str(tensor_key).lower() or 
                                'z*' in str(tensor_key).lower() or
                                'effective' in str(tensor_key).lower()):
                                filtered_items.append(item)
                                
                        elif tensor_type.startswith("High-frequency"):
                            # Look for high-frequency dielectric (ε∞)
                            if ('ε∞' in str(tensor_key) or 
                                'eps_inf' in str(tensor_key).lower() or 
                                'epsilon_inf' in str(tensor_key).lower() or
                                'eps∞' in str(tensor_key) or
                                'Ɛ∞' in str(tensor_key) or
                                'Eps_inf' in str(tensor_key)):
                                filtered_items.append(item)
                                
                        elif tensor_type.startswith("Static"):
                            # Look for static dielectric (ε₀)
                            if ('ε0' in str(tensor_key) or 
                                'eps_0' in str(tensor_key).lower() or 
                                'epsilon_0' in str(tensor_key).lower() or
                                'ε₀' in str(tensor_key) or
                                'Ɛ0' in str(tensor_key) or
                                'Eps_0' in str(tensor_key)):
                                filtered_items.append(item)
                                
                        elif tensor_type.startswith("Refractive"):
                            # For refractive index, use high-frequency dielectric data
                            if ('ε∞' in str(tensor_key) or 
                                'eps_inf' in str(tensor_key).lower() or 
                                'epsilon_inf' in str(tensor_key).lower() or
                                'eps∞' in str(tensor_key) or
                                'Ɛ∞' in str(tensor_key) or
                                'Eps_inf' in str(tensor_key)):
                                filtered_items.append(item)
                
                # Add debug info to the display
                if not filtered_items:
                    # Show what we found and what we were looking for
                    debug_info += f"Looking for: {tensor_type}\n"
                    debug_info += f"No matching items found.\n\n"
                    debug_info += "Sample items:\n"
                    for i, item in enumerate(dielectric_data[:5]):
                        debug_info += f"  [{i}]: {item}\n"
                    
                    # Store debug info for later display
                    self._debug_info = debug_info
                
                return filtered_items if filtered_items else None
            
            elif isinstance(dielectric_data, pd.DataFrame):
                # Filter DataFrame based on tensor type
                if 'Tensor' in dielectric_data.columns:
                    if tensor_type.startswith("Born"):
                        filtered_df = dielectric_data[dielectric_data['Tensor'].str.contains('Born|born|Z\\*|Zstar', case=False, na=False)]
                    elif tensor_type.startswith("High-frequency"):
                        filtered_df = dielectric_data[dielectric_data['Tensor'].str.contains('ε∞|eps_inf|epsilon_inf|Ɛ∞|Eps_inf', case=False, na=False)]
                    elif tensor_type.startswith("Static"):
                        filtered_df = dielectric_data[dielectric_data['Tensor'].str.contains('ε0|eps_0|epsilon_0|ε₀|Ɛ0|Eps_0', case=False, na=False)]
                    elif tensor_type.startswith("Refractive"):
                        # For refractive index, use high-frequency dielectric data
                        filtered_df = dielectric_data[dielectric_data['Tensor'].str.contains('ε∞|eps_inf|epsilon_inf|Ɛ∞|Eps_inf', case=False, na=False)]
                    else:
                        filtered_df = dielectric_data
                    
                    return filtered_df if not filtered_df.empty else None
            
            # For other types, return as-is
            return dielectric_data
            
        except Exception as e:
            return None
    
    def display_dataframe_tensors(self, df, tensor_type):
        """Display tensor data from pandas DataFrame."""
        try:
            if tensor_type.startswith("High-frequency"):
                tensor_rows = df[df['Tensor'] == 'Ɛ∞']
                if tensor_rows.empty:
                    # Try alternative notations
                    tensor_rows = df[df['Tensor'].str.contains('eps_inf|epsilon_inf|ε∞|Eps_inf', case=False, na=False)]
            elif tensor_type.startswith("Static"):
                tensor_rows = df[df['Tensor'] == 'Ɛ0']
                if tensor_rows.empty:
                    # Try alternative notations
                    tensor_rows = df[df['Tensor'].str.contains('eps_0|epsilon_0|ε0|Eps_0', case=False, na=False)]
            else:  # Born charges
                # Try multiple Born charge identifiers
                tensor_rows = df[df['Tensor'].str.contains('Born|born|BORN|Z\\*|Zstar|effective', case=False, na=False)]
            
            if tensor_rows.empty:
                # Debug: show what tensors are available
                available_tensors = df['Tensor'].unique() if 'Tensor' in df.columns else []
                debug_text = f"No {tensor_type} data found.\n\nAvailable tensors in DataFrame:\n"
                for tensor in available_tensors:
                    debug_text += f"  • {tensor}\n"
                debug_text += f"\nDataFrame columns: {list(df.columns)}\n"
                debug_text += f"DataFrame shape: {df.shape}"
                self.tensor_properties_text.setPlainText(debug_text)
                return
            
            # Populate table
            self.dielectric_table.setRowCount(len(tensor_rows))
            
            for i, (_, row) in enumerate(tensor_rows.iterrows()):
                component_item = QTableWidgetItem(str(row.get('Component', '')))
                x_item = QTableWidgetItem(f"{float(row.get('X', 0)):.4f}")
                y_item = QTableWidgetItem(f"{float(row.get('Y', 0)):.4f}")
                z_item = QTableWidgetItem(f"{float(row.get('Z', 0)):.4f}")
                
                self.dielectric_table.setItem(i, 0, component_item)
                self.dielectric_table.setItem(i, 1, x_item)
                self.dielectric_table.setItem(i, 2, y_item)
                self.dielectric_table.setItem(i, 3, z_item)
            
            # Calculate properties
            self.calculate_tensor_properties(tensor_rows, tensor_type)
            
        except Exception as e:
            self.tensor_properties_text.setPlainText(f"Error displaying tensor data: {str(e)}")
    
    def display_dict_tensors(self, tensor_dict, tensor_type):
        """Display tensor data from dictionary format."""
        try:
            # Look for the specific tensor type in the dictionary
            tensor_data = None
            debug_keys = list(tensor_dict.keys())
            
            if tensor_type.startswith("High-frequency"):
                # Try various high-frequency dielectric constant keys
                for key in ['eps_inf', 'epsilon_inf', 'eps∞', 'epsilon∞', 'high_freq_dielectric', 'optical_dielectric']:
                    if key in tensor_dict:
                        tensor_data = tensor_dict[key]
                        break
            elif tensor_type.startswith("Static"):
                # Try various static dielectric constant keys
                for key in ['eps_0', 'epsilon_0', 'eps0', 'epsilon0', 'static_dielectric', 'low_freq_dielectric']:
                    if key in tensor_dict:
                        tensor_data = tensor_dict[key]
                        break
            elif tensor_type.startswith("Born"):
                # Try various Born effective charge keys
                for key in ['born_charges', 'born_effective_charges', 'z_star', 'zstar', 'Z*', 'effective_charges', 
                           'born', 'Born_charges', 'Born_effective_charges', 'Zstar', 'Z_star']:
                    if key in tensor_dict:
                        tensor_data = tensor_dict[key]
                        break
            
            if tensor_data is None:
                # Debug: show what keys are available
                debug_text = f"No {tensor_type} data found in dictionary.\n\n"
                debug_text += f"Available keys in dictionary:\n"
                for key in debug_keys:
                    debug_text += f"  • {key}: {type(tensor_dict[key])}\n"
                debug_text += f"\nSearched for {tensor_type} using these patterns:\n"
                
                if tensor_type.startswith("High-frequency"):
                    debug_text += "  eps_inf, epsilon_inf, eps∞, epsilon∞, high_freq_dielectric, optical_dielectric"
                elif tensor_type.startswith("Static"):
                    debug_text += "  eps_0, epsilon_0, eps0, epsilon0, static_dielectric, low_freq_dielectric"
                elif tensor_type.startswith("Born"):
                    debug_text += "  born_charges, born_effective_charges, z_star, zstar, Z*, effective_charges, born"
                
                self.tensor_properties_text.setPlainText(debug_text)
                return
            
            # Display the tensor data based on format
            if isinstance(tensor_data, (list, tuple)) and len(tensor_data) >= 3:
                # Assume it's diagonal elements [xx, yy, zz]
                self.dielectric_table.setRowCount(3)
                components = ['xx', 'yy', 'zz']
                
                for i, (comp, value) in enumerate(zip(components, tensor_data)):
                    self.dielectric_table.setItem(i, 0, QTableWidgetItem(comp))
                    self.dielectric_table.setItem(i, 1, QTableWidgetItem(f"{float(value):.4f}"))
                    self.dielectric_table.setItem(i, 2, QTableWidgetItem("0.0000"))
                    self.dielectric_table.setItem(i, 3, QTableWidgetItem("0.0000"))
                
                # Calculate properties
                self.calculate_simple_tensor_properties(tensor_data, tensor_type)
                
            elif isinstance(tensor_data, dict):
                # Dictionary with component keys
                self.dielectric_table.setRowCount(len(tensor_data))
                
                for i, (comp, value) in enumerate(tensor_data.items()):
                    self.dielectric_table.setItem(i, 0, QTableWidgetItem(str(comp)))
                    if isinstance(value, (list, tuple)) and len(value) >= 3:
                        self.dielectric_table.setItem(i, 1, QTableWidgetItem(f"{float(value[0]):.4f}"))
                        self.dielectric_table.setItem(i, 2, QTableWidgetItem(f"{float(value[1]):.4f}"))
                        self.dielectric_table.setItem(i, 3, QTableWidgetItem(f"{float(value[2]):.4f}"))
                    else:
                        self.dielectric_table.setItem(i, 1, QTableWidgetItem(f"{float(value):.4f}"))
                        self.dielectric_table.setItem(i, 2, QTableWidgetItem("0.0000"))
                        self.dielectric_table.setItem(i, 3, QTableWidgetItem("0.0000"))
                
                # Calculate properties from dict
                diagonal_values = []
                for comp in ['xx', 'yy', 'zz']:
                    if comp in tensor_data:
                        diagonal_values.append(float(tensor_data[comp]))
                
                if diagonal_values:
                    self.calculate_simple_tensor_properties(diagonal_values, tensor_type)
                else:
                    # Try to show available data
                    properties_text = f"{tensor_type} Dictionary Data:\n\n"
                    for key, value in tensor_data.items():
                        properties_text += f"{key}: {value}\n"
                    self.tensor_properties_text.setPlainText(properties_text)
                
            else:
                # Single value or other format
                properties_text = f"{tensor_type} Data:\n\n"
                properties_text += f"Type: {type(tensor_data)}\n"
                properties_text += f"Value: {tensor_data}\n"
                
                # Try to display in table if it's a matrix-like structure
                if hasattr(tensor_data, 'shape'):
                    properties_text += f"Shape: {tensor_data.shape}\n"
                
                self.tensor_properties_text.setPlainText(properties_text)
                
        except Exception as e:
            self.tensor_properties_text.setPlainText(f"Error displaying dictionary tensor data: {str(e)}")
    
    def display_list_tensors(self, tensor_list, tensor_type):
        """Display tensor data from list format."""
        try:
            # Show the raw data structure first
            debug_text = f"List format tensor data for {tensor_type}:\n\n"
            debug_text += f"List length: {len(tensor_list)}\n"
            debug_text += f"Sample items (first 3):\n"
            
            for i, item in enumerate(tensor_list[:3]):
                debug_text += f"  [{i}]: {item}\n"
            
            self.tensor_properties_text.setPlainText(debug_text)
            
            # Handle different list structures
            if not tensor_list:
                self.tensor_properties_text.setPlainText(f"Empty {tensor_type} list.")
                return
            
            # Check what type of data we have in the list
            first_item = tensor_list[0]
            
            if isinstance(first_item, dict):
                # Check if this is the Born charge format with 'atom' and 'charge' fields
                if 'atom' in first_item and 'charge' in first_item:
                    self.display_born_charge_format(tensor_list, tensor_type)
                # Check if this is the component format with 'Component', 'X', 'Y', 'Z' fields
                elif any(key in first_item for key in ['Component', 'component']) and any(key in first_item for key in ['X', 'Y', 'Z']):
                    self.display_component_format(tensor_list, tensor_type)
                else:
                    # List of dictionaries - try generic component format
                    self.display_dict_list_tensors(tensor_list, tensor_type)
                
            elif isinstance(first_item, (list, tuple)):
                # List of lists/tuples - matrix format
                self.display_matrix_list_tensors(tensor_list, tensor_type)
                
            elif isinstance(first_item, (int, float)):
                # List of numbers - simple format
                self.display_numeric_list_tensors(tensor_list, tensor_type)
                
            else:
                # Unknown format
                properties_text = f"{tensor_type} List Data (Unknown Format):\n\n"
                properties_text += f"List contains items of type: {type(first_item)}\n"
                properties_text += f"First few items:\n"
                for i, item in enumerate(tensor_list[:5]):
                    properties_text += f"  [{i}]: {item}\n"
                self.tensor_properties_text.setPlainText(properties_text)
                
        except Exception as e:
            self.tensor_properties_text.setPlainText(f"Error displaying list tensor data: {str(e)}")
    
    def display_component_format(self, tensor_list, tensor_type):
        """Display tensor data in component format: [{'Component': 'xx', 'X': value, 'Y': value, 'Z': value}, ...]"""
        try:
            import numpy as np
            
            debug_text = f"Processing component format for {tensor_type}:\n\n"
            
            # Group by tensor type if available, otherwise treat as single tensor
            tensor_groups = {}
            
            for item in tensor_list:
                if isinstance(item, dict):
                    # Look for tensor identifier
                    tensor_id = item.get('Tensor', item.get('tensor', 'Main'))
                    component = item.get('Component', item.get('component', ''))
                    
                    if tensor_id not in tensor_groups:
                        tensor_groups[tensor_id] = {}
                    
                    # Extract X, Y, Z values
                    try:
                        x_val = float(item.get('X', 0.0))
                        y_val = float(item.get('Y', 0.0))
                        z_val = float(item.get('Z', 0.0))
                        tensor_groups[tensor_id][component] = [x_val, y_val, z_val]
                    except (ValueError, TypeError):
                        continue
            
            debug_text += f"Found tensor groups: {list(tensor_groups.keys())}\n"
            
            # Calculate total rows needed (4 rows per tensor: 3 matrix + 1 eigenvalue)
            valid_tensors = []
            for tensor_id, components in tensor_groups.items():
                if any(comp in components for comp in ['xx', 'yy', 'zz', 'X', 'Y', 'Z']):
                    valid_tensors.append(tensor_id)
            
            if not valid_tensors:
                debug_text += "No valid tensor data found.\n"
                self.tensor_properties_text.setPlainText(debug_text)
                return
            
            total_rows = len(valid_tensors) * 4
            self.dielectric_table.setRowCount(total_rows)
            debug_text += f"Set table to {total_rows} rows for {len(valid_tensors)} tensors\n\n"
            
            properties_text = f"{tensor_type} - Tensor Format Display:\n"
            properties_text += "=" * 50 + "\n\n"
            
            current_row = 0
            
            for tensor_idx, tensor_id in enumerate(valid_tensors):
                try:
                    components = tensor_groups[tensor_id]
                    debug_text += f"Processing tensor {tensor_idx}: {tensor_id}\n"
                    debug_text += f"  Components: {list(components.keys())}\n"
                    
                    # Build tensor matrix - try different component naming conventions
                    tensor_matrix = None
                    
                    # Try standard xx, yy, zz format
                    if all(comp in components for comp in ['xx', 'yy', 'zz']):
                        tensor_matrix = [
                            components['xx'],
                            components['yy'], 
                            components['zz']
                        ]
                        debug_text += f"  Built matrix from xx/yy/zz components\n"
                    
                    # Try diagonal X, Y, Z format  
                    elif all(comp in components for comp in ['X', 'Y', 'Z']):
                        tensor_matrix = [
                            [components['X'][0], 0, 0],
                            [0, components['Y'][0], 0],
                            [0, 0, components['Z'][0]]
                        ]
                        debug_text += f"  Built diagonal matrix from X/Y/Z components\n"
                    
                    # Try to build full 3x3 matrix if we have enough components
                    elif len(components) >= 6:  # At least 6 components for symmetric matrix
                        tensor_matrix = [[0, 0, 0], [0, 0, 0], [0, 0, 0]]
                        comp_map = {
                            'xx': (0, 0), 'xy': (0, 1), 'xz': (0, 2),
                            'yx': (1, 0), 'yy': (1, 1), 'yz': (1, 2),
                            'zx': (2, 0), 'zy': (2, 1), 'zz': (2, 2)
                        }
                        
                        for comp, (i, j) in comp_map.items():
                            if comp in components:
                                tensor_matrix[i][j] = components[comp][0]  # Use X component
                        
                        debug_text += f"  Built full matrix from available components\n"
                    
                    if tensor_matrix is None:
                        debug_text += f"  Could not build matrix - insufficient components\n"
                        continue
                    
                    # Calculate eigenvalues
                    tensor_array = np.array(tensor_matrix, dtype=float)
                    eigenvalues = np.linalg.eigvals(tensor_array)
                    eigenvalues = np.sort(eigenvalues)[::-1]  # Sort descending
                    
                    # Display tensor name in first row
                    display_name = tensor_id if tensor_id != 'Main' else tensor_type.split()[0]
                    self.dielectric_table.setItem(current_row, 0, QTableWidgetItem(f"{display_name}:"))
                    
                    # Display 3x3 matrix
                    for i in range(3):
                        if i > 0:
                            self.dielectric_table.setItem(current_row + i, 0, QTableWidgetItem(""))
                        
                        # Matrix values
                        for j in range(3):
                            if i < len(tensor_matrix) and j < len(tensor_matrix[i]):
                                value = float(tensor_matrix[i][j])
                                self.dielectric_table.setItem(current_row + i, j + 1, 
                                                            QTableWidgetItem(f"{value:.4f}"))
                    
                    # Eigenvalue row
                    self.dielectric_table.setItem(current_row + 3, 0, QTableWidgetItem("Eig. Value:"))
                    for j, eig_val in enumerate(eigenvalues[:3]):
                        self.dielectric_table.setItem(current_row + 3, j + 1, 
                                                    QTableWidgetItem(f"{float(eig_val):.4f}"))
                    
                    # Add to properties text
                    properties_text += f"{display_name}:\n"
                    properties_text += "Matrix:\n"
                    for i, row in enumerate(tensor_matrix):
                        comp_name = ['xx', 'yy', 'zz'][i] if i < 3 else f'row_{i}'
                        properties_text += f"  {comp_name}: [{row[0]:8.4f} {row[1]:8.4f} {row[2]:8.4f}]\n"
                    properties_text += f"Eigenvalues: [{eigenvalues[0]:.4f}, {eigenvalues[1]:.4f}, {eigenvalues[2]:.4f}]\n"
                    
                    # Calculate refractive indices for dielectric tensors
                    if tensor_type.startswith("High-frequency"):
                        properties_text += "Refractive indices:\n"
                        for i, eig_val in enumerate(eigenvalues):
                            if eig_val > 0:
                                n = np.sqrt(eig_val)
                                properties_text += f"  n{i+1}: {n:.4f}\n"
                        properties_text += f"Average n: {np.sqrt(np.mean([e for e in eigenvalues if e > 0])):.4f}\n"
                        if len([e for e in eigenvalues if e > 0]) >= 2:
                            valid_n = [np.sqrt(e) for e in eigenvalues if e > 0]
                            birefringence = max(valid_n) - min(valid_n)
                            properties_text += f"Birefringence: {birefringence:.4f}\n"
                    
                    elif tensor_type.startswith("Refractive"):
                        # For refractive index, convert dielectric eigenvalues to refractive indices
                        # and display as refractive index tensor
                        properties_text += "Refractive Index Tensor (from ε∞):\n"
                        
                        # Convert tensor matrix from dielectric to refractive index
                        n_matrix = [[0, 0, 0], [0, 0, 0], [0, 0, 0]]
                        for i in range(3):
                            for j in range(3):
                                if tensor_matrix[i][j] > 0:
                                    n_matrix[i][j] = np.sqrt(tensor_matrix[i][j])
                                else:
                                    n_matrix[i][j] = 0
                        
                        # Update table to show refractive indices instead of dielectric
                        for i in range(3):
                            for j in range(3):
                                value = n_matrix[i][j]
                                self.dielectric_table.setItem(current_row - 3 + i, j + 1, 
                                                            QTableWidgetItem(f"{value:.4f}"))
                        
                        # Update eigenvalues to refractive index eigenvalues
                        n_eigenvalues = [np.sqrt(e) if e > 0 else 0 for e in eigenvalues]
                        for j, n_eig in enumerate(n_eigenvalues[:3]):
                            self.dielectric_table.setItem(current_row, j + 1, 
                                                        QTableWidgetItem(f"{n_eig:.4f}"))
                        
                        # Display refractive index matrix in properties
                        properties_text += "Matrix:\n"
                        for i, row in enumerate(n_matrix):
                            comp_name = ['xx', 'yy', 'zz'][i] if i < 3 else f'row_{i}'
                            properties_text += f"  {comp_name}: [{row[0]:8.4f} {row[1]:8.4f} {row[2]:8.4f}]\n"
                        properties_text += f"Principal values: [{n_eigenvalues[0]:.4f}, {n_eigenvalues[1]:.4f}, {n_eigenvalues[2]:.4f}]\n"
                        
                        if len([n for n in n_eigenvalues if n > 0]) >= 2:
                            valid_n = [n for n in n_eigenvalues if n > 0]
                            properties_text += f"Average n: {np.mean(valid_n):.4f}\n"
                            birefringence = max(valid_n) - min(valid_n)
                            properties_text += f"Birefringence (Δn): {birefringence:.4f}\n"
                    current_row += 4

                except Exception as e:
                    error_msg = f"Error processing tensor {tensor_idx}: {str(e)}"
                    properties_text += error_msg + "\n"
                    current_row += 1
            
            self.tensor_properties_text.setPlainText(properties_text)
            
        except ImportError:
            self.tensor_properties_text.setPlainText(f"NumPy required for calculations. Install with: pip install numpy")
        except Exception as e:
            self.tensor_properties_text.setPlainText(f"Error displaying component format: {str(e)}")
    
    def display_born_charge_format(self, tensor_list, tensor_type):
        """Display Born charges in the format: [{'atom': 'Ti', 'charge': [[...]], 'eigenvectors': [...]}]"""
        try:
            import numpy as np
            
            # Calculate total rows needed (4 rows per atom: 3 matrix + 1 eigenvalue)
            valid_atoms = []
            debug_text = f"Processing {len(tensor_list)} items for {tensor_type}:\n"
            
            for i, item in enumerate(tensor_list):
                debug_text += f"Item {i}: {type(item)}\n"
                if isinstance(item, dict):
                    debug_text += f"  Keys: {list(item.keys())}\n"
                    if 'atom' in item:
                        debug_text += f"  Atom: {item['atom']}\n"
                    if 'charge' in item:
                        debug_text += f"  Charge type: {type(item['charge'])}\n"
                        if isinstance(item['charge'], list):
                            debug_text += f"  Charge shape: {len(item['charge'])}x{len(item['charge'][0]) if item['charge'] else 0}\n"
                    
                    if 'atom' in item and 'charge' in item:
                        valid_atoms.append(item)
                        debug_text += f"  -> Added as valid atom\n"
                    else:
                        debug_text += f"  -> Skipped (missing required fields)\n"
                else:
                    debug_text += f"  -> Skipped (not a dict)\n"
            
            debug_text += f"\nTotal valid atoms found: {len(valid_atoms)}\n"
            
            if not valid_atoms:
                self.tensor_properties_text.setPlainText(f"No valid atoms found in {tensor_type} data.\n\n{debug_text}")
                return
            
            total_rows = len(valid_atoms) * 4
            self.dielectric_table.setRowCount(total_rows)
            debug_text += f"Set table to {total_rows} rows\n\n"
            
            # Properties text for detailed information
            properties_text = f"{tensor_type} - WURM Format Display:\n"
            properties_text += "=" * 50 + "\n\n"
            
            current_row = 0
            
            for atom_idx, atom_data in enumerate(valid_atoms):
                try:
                    atom_type = atom_data['atom']
                    charge_matrix = atom_data['charge']
                    
                    debug_text += f"Processing atom {atom_idx}: {atom_type}\n"
                    debug_text += f"  Current row: {current_row}\n"
                    debug_text += f"  Charge matrix: {charge_matrix}\n"
                    
                    # Get eigenvalues if available
                    eigenvalues = None
                    if 'eigenvectors' in atom_data and atom_data['eigenvectors']:
                        # Note: despite the name 'eigenvectors', this contains eigenvalues
                        eig_data = atom_data['eigenvectors']
                        if isinstance(eig_data, list) and len(eig_data) > 0:
                            if isinstance(eig_data[0], list):
                                eigenvalues = eig_data[0]  # Nested list
                            else:
                                eigenvalues = eig_data  # Direct list
                        debug_text += f"  Eigenvalues from data: {eigenvalues}\n"
                    
                    # If no pre-calculated eigenvalues, calculate them
                    if eigenvalues is None:
                        tensor_array = np.array(charge_matrix, dtype=float)
                        eigenvalues = np.linalg.eigvals(tensor_array)
                        eigenvalues = np.sort(eigenvalues)[::-1]  # Sort descending
                        debug_text += f"  Calculated eigenvalues: {eigenvalues}\n"
                    
                    # Display atom type in first row
                    self.dielectric_table.setItem(current_row, 0, QTableWidgetItem(f"{atom_type}:"))
                    debug_text += f"  Set atom name at row {current_row}\n"
                    
                    # Display 3x3 matrix
                    for i in range(3):
                        if i > 0:
                            self.dielectric_table.setItem(current_row + i, 0, QTableWidgetItem(""))
                        
                        # Matrix values
                        for j in range(3):
                            if i < len(charge_matrix) and j < len(charge_matrix[i]):
                                value = float(charge_matrix[i][j])
                                self.dielectric_table.setItem(current_row + i, j + 1, 
                                                            QTableWidgetItem(f"{value:.4f}"))
                                if atom_idx == 1:  # Debug O specifically
                                    debug_text += f"  Set matrix[{i}][{j}] = {value} at table row {current_row + i}, col {j + 1}\n"
                    
                    # Eigenvalue row
                    self.dielectric_table.setItem(current_row + 3, 0, QTableWidgetItem("Eig. Value:"))
                    for j, eig_val in enumerate(eigenvalues[:3]):  # Only show first 3
                        self.dielectric_table.setItem(current_row + 3, j + 1, 
                                                    QTableWidgetItem(f"{float(eig_val):.4f}"))
                    
                    debug_text += f"  Set eigenvalues at row {current_row + 3}\n"
                    
                    # Add to properties text
                    properties_text += f"{atom_type}:\n"
                    properties_text += "Matrix:\n"
                    for i, row in enumerate(charge_matrix):
                        comp_name = ['xx', 'yy', 'zz'][i] if i < 3 else f'row_{i}'
                        properties_text += f"  {comp_name}: [{row[0]:8.4f} {row[1]:8.4f} {row[2]:8.4f}]\n"
                    
                    if eigenvalues is not None and len(eigenvalues) >= 3:
                        properties_text += f"Eigenvalues: [{eigenvalues[0]:.4f}, {eigenvalues[1]:.4f}, {eigenvalues[2]:.4f}]\n\n"
                    else:
                        properties_text += "Eigenvalues: Not available\n\n"
                    
                    current_row += 4
                    debug_text += f"  Advanced to row {current_row}\n\n"
                    
                except Exception as e:
                    error_msg = f"Error processing {atom_data.get('atom', 'unknown')} data: {str(e)}"
                    properties_text += error_msg + "\n"
                    debug_text += error_msg + "\n"
                    current_row += 1
            
            # Update properties display with debug info
            full_text = debug_text + "\n" + properties_text
            self.tensor_properties_text.setPlainText(full_text)
            
        except ImportError:
            self.tensor_properties_text.setPlainText(f"NumPy required for eigenvalue calculations. Install with: pip install numpy")
        except Exception as e:
            self.tensor_properties_text.setPlainText(f"Error displaying Born charge format: {str(e)}")
    
    def display_matrix_list_tensors(self, tensor_list, tensor_type):
        """Display tensor data from list of lists/tuples format."""
        try:
            properties_text = f"{tensor_type} Data (Matrix Format):\n\n"
            
            # Try to interpret as rows of a matrix
            if len(tensor_list) == 3 and all(len(row) >= 3 for row in tensor_list if isinstance(row, (list, tuple))):
                # 3x3 matrix
                self.dielectric_table.setRowCount(3)
                components = ['xx', 'yy', 'zz']
                
                for i, (comp, row) in enumerate(zip(components, tensor_list)):
                    try:
                        x_val = float(row[0])
                        y_val = float(row[1]) if len(row) > 1 else 0.0
                        z_val = float(row[2]) if len(row) > 2 else 0.0
                        
                        self.dielectric_table.setItem(i, 0, QTableWidgetItem(comp))
                        self.dielectric_table.setItem(i, 1, QTableWidgetItem(f"{x_val:.4f}"))
                        self.dielectric_table.setItem(i, 2, QTableWidgetItem(f"{y_val:.4f}"))
                        self.dielectric_table.setItem(i, 3, QTableWidgetItem(f"{z_val:.4f}"))
                        
                        properties_text += f"{comp}: [{x_val:.4f}, {y_val:.4f}, {z_val:.4f}]\n"
                    except (ValueError, TypeError, IndexError):
                        properties_text += f"{comp}: Error parsing row {i}\n"
            else:
                # General matrix format
                self.dielectric_table.setRowCount(min(len(tensor_list), 10))  # Limit to 10 rows
                
                for i, row in enumerate(tensor_list[:10]):
                    try:
                        if isinstance(row, (list, tuple)) and len(row) >= 1:
                            x_val = float(row[0]) if len(row) > 0 else 0.0
                            y_val = float(row[1]) if len(row) > 1 else 0.0
                            z_val = float(row[2]) if len(row) > 2 else 0.0
                            
                            self.dielectric_table.setItem(i, 0, QTableWidgetItem(f"Row_{i+1}"))
                            self.dielectric_table.setItem(i, 1, QTableWidgetItem(f"{x_val:.4f}"))
                            self.dielectric_table.setItem(i, 2, QTableWidgetItem(f"{y_val:.4f}"))
                            self.dielectric_table.setItem(i, 3, QTableWidgetItem(f"{z_val:.4f}"))
                            
                            properties_text += f"Row {i+1}: {row}\n"
                    except (ValueError, TypeError):
                        properties_text += f"Row {i+1}: Error parsing\n"
            
            # Update properties display
            current_text = self.tensor_properties_text.toPlainText()
            self.tensor_properties_text.setPlainText(current_text + "\n\n" + properties_text)
            
        except Exception as e:
            self.tensor_properties_text.setPlainText(f"Error displaying matrix list tensors: {str(e)}")
    
    def display_numeric_list_tensors(self, tensor_list, tensor_type):
        """Display tensor data from list of numbers format."""
        try:
            # Original logic for numeric lists
            if len(tensor_list) == 3:
                # Assume diagonal elements
                self.dielectric_table.setRowCount(3)
                components = ['xx', 'yy', 'zz']
                
                for i, (comp, value) in enumerate(zip(components, tensor_list)):
                    self.dielectric_table.setItem(i, 0, QTableWidgetItem(comp))
                    self.dielectric_table.setItem(i, 1, QTableWidgetItem(f"{float(value):.4f}"))
                    self.dielectric_table.setItem(i, 2, QTableWidgetItem("0.0000"))
                    self.dielectric_table.setItem(i, 3, QTableWidgetItem("0.0000"))
                
                self.calculate_simple_tensor_properties(tensor_list, tensor_type)
                
            elif len(tensor_list) == 9:
                # Assume 3x3 matrix flattened
                self.dielectric_table.setRowCount(3)
                components = ['xx', 'yy', 'zz']
                matrix_indices = [0, 4, 8]  # Diagonal elements
                
                for i, (comp, idx) in enumerate(zip(components, matrix_indices)):
                    self.dielectric_table.setItem(i, 0, QTableWidgetItem(comp))
                    self.dielectric_table.setItem(i, 1, QTableWidgetItem(f"{float(tensor_list[idx]):.4f}"))
                    self.dielectric_table.setItem(i, 2, QTableWidgetItem("0.0000"))
                    self.dielectric_table.setItem(i, 3, QTableWidgetItem("0.0000"))
                
                diagonal_values = [tensor_list[i] for i in matrix_indices]
                self.calculate_simple_tensor_properties(diagonal_values, tensor_type)
            else:
                # Other lengths - show first few values
                self.dielectric_table.setRowCount(min(len(tensor_list), 5))
                
                for i, value in enumerate(tensor_list[:5]):
                    self.dielectric_table.setItem(i, 0, QTableWidgetItem(f"Element_{i+1}"))
                    self.dielectric_table.setItem(i, 1, QTableWidgetItem(f"{float(value):.4f}"))
                    self.dielectric_table.setItem(i, 2, QTableWidgetItem("0.0000"))
                    self.dielectric_table.setItem(i, 3, QTableWidgetItem("0.0000"))
                
                properties_text = f"{tensor_type} Numeric List:\n\n"
                properties_text += f"Length: {len(tensor_list)}\n"
                properties_text += f"Values: {tensor_list}\n"
                
                current_text = self.tensor_properties_text.toPlainText()
                self.tensor_properties_text.setPlainText(current_text + "\n\n" + properties_text)
                
        except Exception as e:
            self.tensor_properties_text.setPlainText(f"Error displaying numeric list tensors: {str(e)}")
    
    def calculate_simple_tensor_properties(self, values, tensor_type):
        """Calculate properties for simple tensor data formats with eigenvalues."""
        try:
            import numpy as np
            
            properties_text = f"{tensor_type} Properties:\n\n"
            
            if len(values) >= 3:
                diagonal_values = [float(v) for v in values[:3]]
                
                properties_text += "Diagonal Elements:\n"
                components = ['xx', 'yy', 'zz']
                for comp, value in zip(components, diagonal_values):
                    properties_text += f"  {comp}: {value:.4f}\n"
                
                # Calculate averages and anisotropy
                avg_value = sum(diagonal_values) / len(diagonal_values)
                max_val = max(diagonal_values)
                min_val = min(diagonal_values)
                anisotropy = (max_val - min_val) / avg_value if avg_value != 0 else 0
                
                properties_text += f"\nAverage value: {avg_value:.4f}\n"
                properties_text += f"Anisotropy: {anisotropy:.4f}\n"
                
                # For diagonal tensors, eigenvalues are the diagonal elements themselves
                eigenvalues = sorted(diagonal_values, reverse=True)
                properties_text += f"\nEigenvalues (diagonal tensor):\n"
                for i, eig_val in enumerate(eigenvalues):
                    properties_text += f"  λ{i+1}: {eig_val:.4f}\n"
                
                # Add eigenvalue row to table
                self.add_eigenvalue_row_to_table(eigenvalues)
                
                # Calculate optical properties
                optical_text = f"Optical Properties for {tensor_type}:\n\n"
                
                if tensor_type.startswith("High-frequency"):
                    # Calculate refractive indices
                    properties_text += "\nRefractive Indices:\n"
                    optical_text += "Refractive Indices:\n"
                    for comp, eps in zip(components, diagonal_values):
                        if eps > 0:
                            n = eps ** 0.5
                            properties_text += f"  n_{comp[0]}: {n:.4f}\n"
                            optical_text += f"  n_{comp[0]}: {n:.4f}\n"
                    
                    # Average refractive index
                    valid_indices = [eps**0.5 for eps in diagonal_values if eps > 0]
                    if valid_indices:
                        avg_n = sum(valid_indices) / len(valid_indices)
                        optical_text += f"\nAverage refractive index: {avg_n:.4f}\n"
                        
                        # Eigenvalue refractive indices (same as diagonal for this case)
                        eig_n = [np.sqrt(eig) for eig in eigenvalues if eig > 0]
                        if eig_n:
                            optical_text += f"Eigenvalue refractive indices: {[f'{n:.4f}' for n in eig_n]}\n"
                        
                        # Birefringence
                        if len(valid_indices) >= 2:
                            birefringence = max(valid_n) - min(valid_n)
                            optical_text += f"Birefringence (Δn): {birefringence:.4f}\n"
                
                elif tensor_type.startswith("Static"):
                    # Static dielectric properties
                    optical_text += f"Static dielectric constant (average): {avg_value:.4f}\n"
                    optical_text += f"Dielectric anisotropy: {anisotropy:.4f}\n"
                    optical_text += f"Eigenvalue dielectric constants: {[f'{eig:.4f}' for eig in eigenvalues]}\n"
                
                # Update optical properties display
                self.optical_properties_text.setPlainText(optical_text)
            
            self.tensor_properties_text.setPlainText(properties_text)
            
        except ImportError:
            self.tensor_properties_text.setPlainText(f"NumPy required for enhanced calculations. Install with: pip install numpy")
        except Exception as e:
            self.tensor_properties_text.setPlainText(f"Error calculating simple properties: {str(e)}")
    
    def calculate_tensor_properties(self, tensor_data, tensor_type):
        """Calculate and display tensor properties from DataFrame with eigenvalues."""
        try:
            import numpy as np
            
            properties_text = f"{tensor_type} Properties:\n\n"
            
            if isinstance(tensor_data, pd.DataFrame):
                # Extract diagonal elements and try to reconstruct full matrix
                diagonal_elements = {}
                full_matrix = None
                
                # Try to extract full 3x3 tensor matrix
                matrix_dict = {}
                for _, row in tensor_data.iterrows():
                    comp = str(row.get('Component', ''))
                    x_val = float(row.get('X', 0))
                    y_val = float(row.get('Y', 0))
                    z_val = float(row.get('Z', 0))
                    
                    matrix_dict[comp] = [x_val, y_val, z_val]
                    
                    if comp in ['xx', 'yy', 'zz']:
                        diagonal_elements[comp] = x_val
                
                # Try to build full 3x3 matrix
                component_order = ['xx', 'xy', 'xz', 'yx', 'yy', 'yz', 'zx', 'zy', 'zz']
                if all(comp in matrix_dict for comp in ['xx', 'yy', 'zz']):
                    # At minimum we have diagonal elements
                    full_matrix = np.zeros((3, 3))
                    
                    # Fill in what we have
                    for comp, values in matrix_dict.items():
                        if len(comp) == 2:
                            row_map = {'x': 0, 'y': 1, 'z': 2}
                            col_map = {'x': 0, 'y': 1, 'z': 2}
                            
                            if comp[0] in row_map and comp[1] in col_map:
                                i = row_map[comp[0]]
                                j = col_map[comp[1]]
                                full_matrix[i, j] = values[0]  # Use X component for tensor value
                
                if diagonal_elements:
                    properties_text += "Diagonal Elements:\n"
                    for comp, value in diagonal_elements.items():
                        properties_text += f"  {comp}: {value:.4f}\n"
                    
                    # Calculate averages and anisotropy
                    values = list(diagonal_elements.values())
                    if len(values) >= 2:
                        avg_value = sum(values) / len(values)
                        max_val = max(values)
                        min_val = min(values)
                        anisotropy = (max_val - min_val) / avg_value if avg_value != 0 else 0
                        
                        properties_text += f"\nAverage value: {avg_value:.4f}\n"
                        properties_text += f"Anisotropy: {anisotropy:.4f}\n"
                        
                        # Calculate eigenvalues if we have a full matrix
                        if full_matrix is not None and np.any(full_matrix):
                            try:
                                eigenvalues = np.linalg.eigvals(full_matrix)
                                eigenvalues = np.sort(eigenvalues)[::-1]  # Sort descending
                                properties_text += f"\nEigenvalues:\n"
                                for i, eig_val in enumerate(eigenvalues):
                                    properties_text += f"  λ{i+1}: {eig_val:.4f}\n"
                                
                                # Update table to show eigenvalues
                                self.add_eigenvalue_row_to_table(eigenvalues)
                                
                            except Exception as e:
                                properties_text += f"\nEigenvalue calculation failed: {str(e)}\n"
                        
                        # Calculate optical properties
                        optical_text = f"Optical Properties for {tensor_type}:\n\n"
                        
                        if tensor_type.startswith("High-frequency"):
                            # Calculate refractive indices
                            properties_text += "\nRefractive Indices:\n"
                            optical_text += "Refractive Indices:\n"
                            for comp, eps in diagonal_elements.items():
                                if eps > 0:
                                    n = eps ** 0.5
                                    properties_text += f"  n_{comp[0]}: {n:.4f}\n"
                                    optical_text += f"  n_{comp[0]}: {n:.4f}\n"
                            
                            # Average refractive index
                            valid_indices = [eps**0.5 for eps in values if eps > 0]
                            if valid_indices:
                                avg_n = sum(valid_indices) / len(valid_indices)
                                optical_text += f"\nAverage refractive index: {avg_n:.4f}\n"
                                
                                # Eigenvalue-based refractive indices
                                if full_matrix is not None and np.any(full_matrix):
                                    eigenvalues = np.linalg.eigvals(full_matrix)
                                    eig_n = [np.sqrt(eig) for eig in eigenvalues if eig > 0]
                                    if eig_n:
                                        optical_text += f"Eigenvalue refractive indices: {[f'{n:.4f}' for n in sorted(eig_n, reverse=True)]}\n"
                                
                                # Birefringence
                                if len(valid_indices) >= 2:
                                    birefringence = max(valid_indices) - min(valid_indices)
                                    optical_text += f"Birefringence (Δn): {birefringence:.4f}\n"
                        
                        elif tensor_type.startswith("Static"):
                            # Static dielectric properties
                            optical_text += f"Static dielectric constant (average): {avg_value:.4f}\n"
                            optical_text += f"Dielectric anisotropy: {anisotropy:.4f}\n"
                            
                            if full_matrix is not None and np.any(full_matrix):
                                eigenvalues = np.linalg.eigvals(full_matrix)
                                optical_text += f"Eigenvalue dielectric constants: {[f'{eig:.4f}' for eig in sorted(eigenvalues, reverse=True)]}\n"
                        
                        # Update optical properties display
                        self.optical_properties_text.setPlainText(optical_text)
            
            self.tensor_properties_text.setPlainText(properties_text)
            
        except ImportError:
            self.tensor_properties_text.setPlainText(f"NumPy required for eigenvalue calculations. Install with: pip install numpy")
        except Exception as e:
            self.tensor_properties_text.setPlainText(f"Error calculating properties: {str(e)}")
    
    def add_eigenvalue_row_to_table(self, eigenvalues):
        """Add eigenvalue row to the dielectric table."""
        try:
            current_rows = self.dielectric_table.rowCount()
            self.dielectric_table.setRowCount(current_rows + 1)
            
            # Add eigenvalue label
            self.dielectric_table.setItem(current_rows, 0, QTableWidgetItem("Eig. Value:"))
            
            # Add eigenvalue values
            for i, eig_val in enumerate(eigenvalues[:3]):  # Only show first 3
                self.dielectric_table.setItem(current_rows, i + 1, QTableWidgetItem(f"{eig_val:.4f}"))
                
        except Exception as e:
            pass  # Silently fail if table update doesn't work
    
    def update_advanced_display(self):
        """Update the advanced properties display."""
        self.phonon_analysis_text.clear()
        self.structure_info_text.clear()
        
        if not self.current_mineral or self.current_mineral not in self.database:
            return
        
        mineral_data = self.database[self.current_mineral]
        
        # Phonon mode analysis
        phonon_text = f"Phonon Mode Analysis for {self.current_mineral}\n"
        phonon_text += "=" * 50 + "\n\n"
        
        # Use the new method that handles phonon_modes conversion
        modes = self.get_modes_for_mineral(self.current_mineral)
        if modes:
            # Analyze mode distribution
            raman_active = []
            ir_active = []
            silent_modes = []
            
            for mode in modes:
                if len(mode) >= 2:
                    pos, sym = mode[0], mode[1]
                    if self.is_raman_active(sym):
                        raman_active.append(pos)
                    elif 'u' in str(sym).lower():
                        ir_active.append(pos)
                    else:
                        silent_modes.append(pos)
            
            phonon_text += f"Mode Distribution:\n"
            phonon_text += f"  Raman Active: {len(raman_active)} modes\n"
            phonon_text += f"  IR Active: {len(ir_active)} modes\n"
            phonon_text += f"  Silent: {len(silent_modes)} modes\n\n"
            
            if raman_active:
                phonon_text += f"Raman Active Frequencies:\n"
                raman_active.sort()
                for i, freq in enumerate(raman_active):
                    if i % 5 == 0 and i > 0:
                        phonon_text += "\n"
                    phonon_text += f"{freq:6.1f} "
                phonon_text += "\n\n"
            
            # Frequency ranges
            all_freqs = [mode[0] for mode in modes if len(mode) >= 1]
            if all_freqs:
                phonon_text += f"Frequency Range: {min(all_freqs):.1f} - {max(all_freqs):.1f} cm⁻¹\n"
                
                # Common frequency regions
                low_freq = [f for f in all_freqs if f < 400]
                mid_freq = [f for f in all_freqs if 400 <= f < 1000]
                high_freq = [f for f in all_freqs if f >= 1000]
                
                phonon_text += f"Low frequency (<400 cm⁻¹): {len(low_freq)} modes\n"
                phonon_text += f"Mid frequency (400-1000 cm⁻¹): {len(mid_freq)} modes\n"
                phonon_text += f"High frequency (>1000 cm⁻¹): {len(high_freq)} modes\n"
        
        self.phonon_analysis_text.setPlainText(phonon_text)
        
        # Crystal structure information
        structure_text = f"Crystal Structure Information for {self.current_mineral}\n"
        structure_text += "=" * 50 + "\n\n"
        
        structure_text += f"Crystal System: {mineral_data.get('crystal_system', 'Unknown')}\n"
        structure_text += f"Point Group: {mineral_data.get('point_group', 'Unknown')}\n"
        structure_text += f"Space Group: {mineral_data.get('space_group', 'Unknown')}\n\n"
        
        # Additional properties
        additional_props = []
        for key in ['chemical_formula', 'density', 'hardness', 'color']:
            if key in mineral_data and mineral_data[key]:
                prop_name = key.replace('_', ' ').title()
                additional_props.append(f"{prop_name}: {mineral_data[key]}")
        
        if additional_props:
            structure_text += "Additional Properties:\n"
            for prop in additional_props:
                structure_text += f"  {prop}\n"
        
        self.structure_info_text.setPlainText(structure_text)
    
    def update_database_stats(self):
        """Update database statistics display."""
        if not hasattr(self, 'stats_text'):
            return
        
        total_minerals = len([k for k in self.database.keys() if not k.startswith('__') and isinstance(self.database[k], dict)])
        
        if total_minerals == 0:
            self.stats_text.setPlainText("Database is empty.")
            return
        
        # Calculate statistics
        total_modes = 0
        total_raman_modes = 0
        minerals_with_dielectric = 0
        minerals_with_phonon_data = 0
        minerals_with_original_modes = 0
        crystal_systems = {}
        
        for mineral_name, mineral_data in self.database.items():
            if mineral_name.startswith('__'):
                continue
            
            # Skip non-dictionary values (metadata keys)
            if not isinstance(mineral_data, dict):
                continue

            # Use the new method that handles phonon_modes conversion
            modes = self.get_modes_for_mineral(mineral_name)
            total_modes += len(modes)
            
            # Check if this mineral originally had modes vs phonon_modes
            if 'modes' in mineral_data and mineral_data['modes']:
                minerals_with_original_modes += 1
            elif 'phonon_modes' in mineral_data and mineral_data['phonon_modes'] is not None:
                minerals_with_phonon_data += 1
            
            # Count Raman active modes
            for mode in modes:
                if len(mode) >= 2 and self.is_raman_active(mode[1]):
                    total_raman_modes += 1
            
            # Check for dielectric data
            if self.has_dielectric_data(mineral_data):
                minerals_with_dielectric += 1
            
            # Count crystal systems
            crystal_system = mineral_data.get('crystal_system', 'Unknown')
            crystal_systems[crystal_system] = crystal_systems.get(crystal_system, 0) + 1
        
        # Format statistics
        stats_text = f"Database Statistics:\n\n"
        stats_text += f"Total Minerals: {total_minerals}\n"
        stats_text += f"Total Modes: {total_modes}\n"
        stats_text += f"Raman Active Modes: {total_raman_modes}\n"
        stats_text += f"With Dielectric Data: {minerals_with_dielectric}\n\n"
        
        # Mode data sources
        stats_text += f"Mode Data Sources:\n"
        stats_text += f"  Original 'modes': {minerals_with_original_modes}\n"
        stats_text += f"  Converted 'phonon_modes': {minerals_with_phonon_data}\n"
        stats_text += f"  Total with modes: {minerals_with_original_modes + minerals_with_phonon_data}\n\n"
        
        stats_text += "Crystal Systems:\n"
        for system, count in sorted(crystal_systems.items()):
            stats_text += f"  {system}: {count}\n"
        
        if total_minerals > 0:
            avg_modes = total_modes / total_minerals
            stats_text += f"\nAverage modes per mineral: {avg_modes:.1f}"
        
        self.stats_text.setPlainText(stats_text)
    
    def view_detailed_mineral_info(self):
        """Show detailed mineral information in a popup window."""
        if not self.current_mineral:
            QMessageBox.warning(self, "No Selection", "Please select a mineral first.")
            return
        
        mineral_data = self.database[self.current_mineral]
        
        # Create detailed info dialog
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Detailed Information: {self.current_mineral}")
        dialog.setMinimumSize(600, 500)
        
        layout = QVBoxLayout(dialog)
        
        # Create tabbed display
        tab_widget = QTabWidget()
        
        # General info tab
        general_tab = QWidget()
        general_layout = QVBoxLayout(general_tab)
        general_text = QTextEdit()
        general_text.setReadOnly(True)
        
        info_content = f"Mineral: {self.current_mineral}\n\n"
        for key, value in mineral_data.items():
            if key != 'modes' and not key.startswith('_'):
                info_content += f"{key.replace('_', ' ').title()}: {value}\n"
        
        general_text.setPlainText(info_content)
        general_layout.addWidget(general_text)
        tab_widget.addTab(general_tab, "General")
        
        # Modes tab
        modes_tab = QWidget()
        modes_layout = QVBoxLayout(modes_tab)
        modes_text = QTextEdit()
        modes_text.setReadOnly(True)
        
        modes_content = f"Raman Modes for {self.current_mineral}\n\n"
        modes = mineral_data.get('modes', [])
        for i, mode in enumerate(modes):
            if len(mode) >= 3:
                pos, sym, intensity = mode[:3]
                active = "Yes" if self.is_raman_active(sym) else "No"
                modes_content += f"{i+1:2d}. {pos:7.1f} cm⁻¹  {sym:>6}  {intensity:6.3f}  Raman: {active}\n"
        
        modes_text.setPlainText(modes_content)
        modes_layout.addWidget(modes_text)
        tab_widget.addTab(modes_tab, "Modes")
        
        layout.addWidget(tab_widget)
        
        # Close button
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dialog.close)
        layout.addWidget(close_btn)
        
        dialog.exec()
    
    def view_full_dielectric_data(self):
        """Show full dielectric tensor data in a popup window."""
        if not self.current_mineral:
            QMessageBox.warning(self, "No Selection", "Please select a mineral first.")
            return
        
        mineral_data = self.database[self.current_mineral]
        
        # Look for dielectric tensor data - fix DataFrame boolean evaluation
        dielectric_data = None
        if 'dielectric_tensors' in mineral_data:
            data = mineral_data['dielectric_tensors']
            if data is not None and not (isinstance(data, pd.DataFrame) and data.empty):
                dielectric_data = data
        
        if dielectric_data is None and 'dielectric_tensor' in mineral_data:
            data = mineral_data['dielectric_tensor']
            if data is not None and not (isinstance(data, pd.DataFrame) and data.empty):
                dielectric_data = data
        
        if dielectric_data is None:
            QMessageBox.information(self, "No Data", "No dielectric tensor data available for this mineral.")
            return
        
        # Create dielectric data dialog
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Dielectric Tensor Data: {self.current_mineral}")
        dialog.setMinimumSize(700, 600)
        
        layout = QVBoxLayout(dialog)
        
        # Display the data
        data_text = QTextEdit()
        data_text.setReadOnly(True)
        
        if isinstance(dielectric_data, pd.DataFrame):
            data_content = f"Dielectric Tensor Data for {self.current_mineral}\n\n"
            data_content += dielectric_data.to_string()
        else:
            data_content = f"Dielectric Tensor Data for {self.current_mineral}\n\n"
            data_content += str(dielectric_data)
        
        data_text.setPlainText(data_content)
        layout.addWidget(data_text)
        
        # Close button
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dialog.close)
        layout.addWidget(close_btn)
        
        dialog.exec()
    
    def export_calculated_spectrum(self):
        """Export the calculated spectrum to a file."""
        if not self.current_mineral:
            QMessageBox.warning(self, "No Selection", "Please select a mineral first.")
            return
        
        # Use the new method that handles phonon_modes conversion
        modes = self.get_modes_for_mineral(self.current_mineral)
        
        if not modes:
            QMessageBox.warning(self, "No Data", "No modes available for export.")
            return
        
        # Get file path
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Calculated Spectrum",
            f"{self.current_mineral}_calculated_spectrum.txt",
            "Text files (*.txt);;CSV files (*.csv);;All files (*.*)"
        )
        
        if not file_path:
            return
        
        try:
            # Create calculated spectrum data
            positions = []
            intensities = []
            symmetries = []
            
            for mode in modes:
                if len(mode) >= 3:
                    pos, sym, intensity = mode[:3]
                    
                    # Include only Raman active modes if filter is on
                    if self.raman_only_check.isChecked():
                        if not self.is_raman_active(sym):
                            continue
                    
                    positions.append(float(pos))
                    intensities.append(float(intensity))
                    symmetries.append(str(sym))
            
            if not positions:
                QMessageBox.warning(self, "No Data", "No modes available for export after filtering.")
                return
            
            # Write data
            with open(file_path, 'w') as f:
                f.write(f"# Calculated Raman Spectrum for {self.current_mineral}\n")
                f.write(f"# Generated by RamanLab Mineral Modes Database Browser\n")
                f.write(f"# Total modes: {len(positions)}\n")
                f.write("# Wavenumber(cm-1)\tIntensity\tSymmetry\tRaman_Active\n")
                
                for pos, intensity, sym in zip(positions, intensities, symmetries):
                    active = "Yes" if self.is_raman_active(sym) else "No"
                    f.write(f"{pos:.2f}\t{intensity:.6f}\t{sym}\t{active}\n")
            
            QMessageBox.information(self, "Success", f"Spectrum exported to:\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export spectrum:\n{str(e)}")
    
    def inspect_database_structure(self):
        """Inspect and display database structure for debugging."""
        dialog = QDialog(self)
        dialog.setWindowTitle("Database Structure Inspector")
        dialog.setMinimumSize(800, 600)
        
        layout = QVBoxLayout(dialog)
        
        # Create tabs for different inspections
        tab_widget = QTabWidget()
        
        # Overview tab
        overview_tab = QWidget()
        overview_layout = QVBoxLayout(overview_tab)
        overview_text = QTextEdit()
        overview_text.setReadOnly(True)
        
        # Generate overview
        overview_content = "Database Structure Overview\n" + "=" * 50 + "\n\n"
        overview_content += f"Total items in database: {len(self.database)}\n\n"
        
        overview_content += "Database keys and types:\n"
        for key, value in self.database.items():
            overview_content += f"  {key}: {type(value)}\n"
            if isinstance(value, dict):
                overview_content += f"    Dict keys: {list(value.keys())}\n"
        
        overview_text.setPlainText(overview_content)
        overview_layout.addWidget(overview_text)
        tab_widget.addTab(overview_tab, "Overview")
        
        # Sample mineral tab
        sample_tab = QWidget()
        sample_layout = QVBoxLayout(sample_tab)
        sample_text = QTextEdit()
        sample_text.setReadOnly(True)
        
        # Generate sample mineral details
        sample_content = "Sample Mineral Data Structure\n" + "=" * 50 + "\n\n"
        
        # Find a mineral with data
        sample_mineral = None
        for name, data in self.database.items():
            if not name.startswith('__') and isinstance(data, dict) and data:
                sample_mineral = name
                break
        
        if sample_mineral:
            sample_data = self.database[sample_mineral]
            sample_content += f"Sample mineral: {sample_mineral}\n\n"
            sample_content += "Complete data structure:\n"
            sample_content += self._format_data_structure(sample_data, indent=0)
        else:
            sample_content += "No suitable sample mineral found.\n"
        
        sample_text.setPlainText(sample_content)
        sample_layout.addWidget(sample_text)
        tab_widget.addTab(sample_tab, "Sample Mineral")
        
        # Dielectric data tab
        dielectric_tab = QWidget()
        dielectric_layout = QVBoxLayout(dielectric_tab)
        dielectric_text = QTextEdit()
        dielectric_text.setReadOnly(True)
        
        # Generate dielectric data analysis
        dielectric_content = "Dielectric Tensor Data Analysis\n" + "=" * 50 + "\n\n"
        
        minerals_with_dielectric = []
        dielectric_keys_found = set()
        
        for name, data in self.database.items():
            if not name.startswith('__') and isinstance(data, dict):
                has_dielectric = False
                for key in data.keys():
                    if any(term in key.lower() for term in ['dielectric', 'tensor', 'born', 'eps', 'epsilon']):
                        has_dielectric = True
                        dielectric_keys_found.add(key)
                
                if has_dielectric:
                    minerals_with_dielectric.append(name)
        
        dielectric_content += f"Minerals with potential dielectric data: {len(minerals_with_dielectric)}\n\n"
        dielectric_content += "Dielectric-related keys found:\n"
        for key in sorted(dielectric_keys_found):
            dielectric_content += f"  • {key}\n"
        
        dielectric_content += "\nDetailed dielectric data for first few minerals:\n\n"
        
        for i, mineral_name in enumerate(minerals_with_dielectric[:3]):
            mineral_data = self.database[mineral_name]
            dielectric_content += f"{i+1}. {mineral_name}:\n"
            
            for key, value in mineral_data.items():
                if any(term in key.lower() for term in ['dielectric', 'tensor', 'born', 'eps', 'epsilon']):
                    dielectric_content += f"   {key}: {type(value)}\n"
                    if isinstance(value, pd.DataFrame):
                        dielectric_content += f"     DataFrame shape: {value.shape}\n"
                        if 'Tensor' in value.columns:
                            unique_tensors = value['Tensor'].unique()
                            dielectric_content += f"     Tensor types: {list(unique_tensors)}\n"
                    elif isinstance(value, dict):
                        dielectric_content += f"     Dict keys: {list(value.keys())}\n"
            dielectric_content += "\n"
        
        dielectric_text.setPlainText(dielectric_content)
        dielectric_layout.addWidget(dielectric_text)
        tab_widget.addTab(dielectric_tab, "Dielectric Data")
        
        layout.addWidget(tab_widget)
        
        # Close button
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dialog.close)
        layout.addWidget(close_btn)
        
        dialog.exec()
    
    def _format_data_structure(self, data, indent=0, max_depth=3):
        """Helper method to format data structure for display."""
        if indent > max_depth:
            return "  " * indent + "... (max depth reached)\n"
        
        result = ""
        indent_str = "  " * indent
        
        if isinstance(data, dict):
            for key, value in data.items():
                result += f"{indent_str}{key}: {type(value)}\n"
                if isinstance(value, (dict, list, pd.DataFrame)) and indent < max_depth:
                    if isinstance(value, pd.DataFrame):
                        result += f"{indent_str}  Shape: {value.shape}\n"
                        result += f"{indent_str}  Columns: {list(value.columns)}\n"
                        if not value.empty:
                            result += f"{indent_str}  Sample data:\n{indent_str}    {value.head(2).to_string()}\n"
                    elif isinstance(value, list) and len(value) > 0:
                        result += f"{indent_str}  Length: {len(value)}\n"
                        result += f"{indent_str}  Sample items: {value[:3]}...\n"
                    elif isinstance(value, dict):
                        result += self._format_data_structure(value, indent + 1, max_depth)
        elif isinstance(data, list):
            result += f"{indent_str}List length: {len(data)}\n"
            if data and indent < max_depth:
                result += f"{indent_str}Sample items:\n"
                for i, item in enumerate(data[:3]):
                    result += f"{indent_str}  [{i}]: {type(item)} = {item}\n"
        else:
            result += f"{indent_str}{type(data)}: {data}\n"
        
        return result
    
    def launch_mineral_comparison(self):
        """Launch mineral comparison tool."""
        QMessageBox.information(
            self,
            "Mineral Comparison",
            "Mineral comparison tool will be implemented.\n\n"
            "This will allow side-by-side comparison of calculated spectra,\n"
            "dielectric properties, and structural parameters."
        )
    
    def launch_batch_analysis(self):
        """Launch batch analysis tool."""
        QMessageBox.information(
            self,
            "Batch Analysis",
            "Batch analysis tool will be implemented.\n\n"
            "This will provide:\n"
            "• Bulk spectrum calculations\n"
            "• Statistical analysis across minerals\n"
            "• Export capabilities for multiple minerals"
        )
    
    def display_dict_list_tensors(self, tensor_list, tensor_type):
        """Display tensor data from list of dictionaries format with full 3x3 matrices and eigenvalues."""
        try:
            # Handle both formats:
            # Format 1: [{'Atom': 'Ti', 'Component': 'xx', 'X': 6.0032, 'Y': -0.0005, 'Z': 0.0003}, ...]
            # Format 2: [{'Tensor': 'ε∞', 'Component': 'xx', 'X': '5.9405', 'Y': '0.0000', 'Z': '0.0000'}, ...]
            
            atom_tensors = {}
            
            for item_dict in tensor_list:
                if not isinstance(item_dict, dict):
                    continue
                
                # Extract identifier (atom type or tensor type)
                identifier = (item_dict.get('Atom', item_dict.get('atom')) or 
                            item_dict.get('Tensor', item_dict.get('tensor', 'Unknown')))
                
                # Extract component (xx, yy, zz)
                component = item_dict.get('Component', item_dict.get('component', ''))
                
                # Extract X, Y, Z values for this component (handle both string and numeric)
                try:
                    x_val = float(item_dict.get('X', 0.0))
                    y_val = float(item_dict.get('Y', 0.0))
                    z_val = float(item_dict.get('Z', 0.0))
                except (ValueError, TypeError):
                    continue  # Skip invalid entries
                
                if identifier and component:
                    if identifier not in atom_tensors:
                        atom_tensors[identifier] = {}
                    
                    # Store the component row
                    atom_tensors[identifier][component] = [x_val, y_val, z_val]
            
            if not atom_tensors:
                # Show debug info
                debug_text = f"No valid {tensor_type} tensor data found.\n\n"
                debug_text += "Sample data structure:\n"
                for i, item in enumerate(tensor_list[:3]):
                    debug_text += f"  [{i}]: {item}\n"
                self.tensor_properties_text.setPlainText(debug_text)
                return
            
            # Convert to 3x3 matrices and display
            self.display_wurm_format_from_components(atom_tensors, tensor_type)
            
        except Exception as e:
            self.tensor_properties_text.setPlainText(f"Error displaying dict list tensors: {str(e)}")
    
    def display_wurm_format_from_components(self, atom_tensors, tensor_type):
        """Display tensors in WURM.info format from component data."""
        try:
            import numpy as np
            
            # Calculate total rows needed
            total_rows = 0
            valid_atoms = []
            
            for atom_type, components in atom_tensors.items():
                # Check if we have all three components (xx, yy, zz)
                if all(comp in components for comp in ['xx', 'yy', 'zz']):
                    valid_atoms.append(atom_type)
                    total_rows += 4  # 3 matrix rows + 1 eigenvalue row
            
            if not valid_atoms:
                self.tensor_properties_text.setPlainText(f"No complete 3x3 {tensor_type} tensors found.")
                return
            
            self.dielectric_table.setRowCount(total_rows)
            
            # Properties text for detailed information
            properties_text = f"{tensor_type} - WURM Format Display:\n"
            properties_text += "=" * 50 + "\n\n"
            
            current_row = 0
            
            for atom_type in sorted(valid_atoms):  # Sort for consistent display
                components = atom_tensors[atom_type]
                
                try:
                    # Build 3x3 matrix from components
                    tensor_matrix = [
                        components['xx'],  # [xx_x, xx_y, xx_z]
                        components['yy'],  # [yy_x, yy_y, yy_z] 
                        components['zz']   # [zz_x, zz_y, zz_z]
                    ]
                    
                    # Convert to numpy array for eigenvalue calculation
                    tensor_array = np.array(tensor_matrix, dtype=float)
                    
                    # Calculate eigenvalues
                    eigenvalues = np.linalg.eigvals(tensor_array)
                    eigenvalues = np.sort(eigenvalues)[::-1]  # Sort descending
                    
                    # Display atom type in first row
                    self.dielectric_table.setItem(current_row, 0, QTableWidgetItem(f"{atom_type}:"))
                    
                    # Display 3x3 matrix
                    for i in range(3):
                        if i > 0:
                            self.dielectric_table.setItem(current_row + i, 0, QTableWidgetItem(""))
                        
                        # Matrix values
                        for j in range(3):
                            value = float(tensor_matrix[i][j])
                            self.dielectric_table.setItem(current_row + i, j + 1, 
                                                        QTableWidgetItem(f"{value:.4f}"))
                    
                    # Eigenvalue row
                    self.dielectric_table.setItem(current_row + 3, 0, QTableWidgetItem("Eig. Value:"))
                    for j, eig_val in enumerate(eigenvalues):
                        if j < 3:
                            self.dielectric_table.setItem(current_row + 3, j + 1, 
                                                        QTableWidgetItem(f"{eig_val:.4f}"))
                    
                    # Add to properties text
                    properties_text += f"{atom_type}:\n"
                    properties_text += "Matrix:\n"
                    for i, row in enumerate(tensor_matrix):
                        comp_name = ['xx', 'yy', 'zz'][i]
                        properties_text += f"  {comp_name}: [{row[0]:8.4f} {row[1]:8.4f} {row[2]:8.4f}]\n"
                    properties_text += f"Eigenvalues: [{eigenvalues[0]:.4f}, {eigenvalues[1]:.4f}, {eigenvalues[2]:.4f}]\n\n"
                    
                    current_row += 4
                    
                except Exception as e:
                    error_msg = f"Error processing {atom_type} tensor: {str(e)}"
                    properties_text += error_msg + "\n"
                    current_row += 1
            
            # Update properties display
            current_text = self.tensor_properties_text.toPlainText()
            self.tensor_properties_text.setPlainText(current_text + "\n\n" + properties_text)
            
        except ImportError:
            self.tensor_properties_text.setPlainText(f"NumPy required for eigenvalue calculations. Install with: pip install numpy")
        except Exception as e:
            self.tensor_properties_text.setPlainText(f"Error displaying WURM format tensors: {str(e)}")
    
    def convert_phonon_modes_to_modes(self, phonon_modes_df):
        """Convert phonon_modes DataFrame to the expected modes list format."""
        try:
            if not isinstance(phonon_modes_df, pd.DataFrame):
                return []
            
            modes = []
            
            for _, row in phonon_modes_df.iterrows():
                # Extract frequency - try TO_Frequency first, then other frequency columns
                frequency = None
                for freq_col in ['TO_Frequency', 'Frequency', 'freq', 'wavenumber']:
                    if freq_col in row and pd.notna(row[freq_col]) and str(row[freq_col]).strip():
                        try:
                            frequency = float(row[freq_col])
                            break
                        except (ValueError, TypeError):
                            continue
                
                if frequency is None:
                    continue
                
                # Extract symmetry/activity
                symmetry = str(row.get('Activity', '')).strip()
                if not symmetry:
                    symmetry = str(row.get('Symmetry', '')).strip()
                if not symmetry:
                    symmetry = 'unknown'
                
                # Skip acoustic modes
                if symmetry.lower() in ['ac', 'acoustic', '']:
                    continue
                
                # Extract intensity - try different intensity columns
                intensity = 1.0  # Default intensity
                for int_col in ['I_Total', 'Intensity', 'intensity', 'I_Parallel', 'I_Perpendicular']:
                    if int_col in row and pd.notna(row[int_col]) and str(row[int_col]).strip():
                        try:
                            intensity_val = float(row[int_col])
                            if intensity_val > 0:
                                intensity = intensity_val
                                break
                        except (ValueError, TypeError):
                            continue
                
                # Normalize very large intensities (common in calculated data)
                if intensity > 1e10:
                    intensity = intensity / 1e38  # Scale down very large values
                
                # Create mode tuple (frequency, symmetry, intensity)
                mode = (frequency, symmetry, intensity)
                modes.append(mode)
            
            return modes
            
        except Exception as e:
            print(f"Error converting phonon_modes: {e}")
            return []
    
    def get_modes_for_mineral(self, mineral_name):
        """Get modes for a mineral, converting from phonon_modes if necessary."""
        if mineral_name not in self.database:
            return []
        
        mineral_data = self.database[mineral_name]
        
        # First try to get existing modes
        if 'modes' in mineral_data and mineral_data['modes']:
            return mineral_data['modes']
        
        # If no modes, try to convert from phonon_modes
        if 'phonon_modes' in mineral_data and mineral_data['phonon_modes'] is not None:
            phonon_modes = mineral_data['phonon_modes']
            converted_modes = self.convert_phonon_modes_to_modes(phonon_modes)
            
            # Cache the converted modes in the mineral data
            if converted_modes:
                mineral_data['modes'] = converted_modes
                return converted_modes
        
        return []


class MineralEditDialog(QDialog):
    """Dialog for editing mineral properties."""
    
    def __init__(self, parent=None, mineral_data=None):
        super().__init__(parent)
        self.mineral_data = mineral_data or {}
        
        self.setWindowTitle("Edit Mineral" if mineral_data else "Add Mineral")
        self.setMinimumSize(400, 300)
        
        self.setup_ui()
    
    def setup_ui(self):
        """Set up the dialog UI."""
        layout = QVBoxLayout(self)
        
        # Form layout for mineral properties
        form_layout = QFormLayout()
        
        # Name
        self.name_edit = QLineEdit()
        self.name_edit.setText(self.mineral_data.get('name', ''))
        form_layout.addRow("Name:", self.name_edit)
        
        # Crystal system
        self.crystal_system_combo = QComboBox()
        crystal_systems = [
            "Cubic", "Tetragonal", "Orthorhombic", "Hexagonal", 
            "Trigonal", "Monoclinic", "Triclinic"
        ]
        self.crystal_system_combo.addItems(crystal_systems)
        current_crystal = self.mineral_data.get('crystal_system', '')
        if current_crystal in crystal_systems:
            self.crystal_system_combo.setCurrentText(current_crystal)
        form_layout.addRow("Crystal System:", self.crystal_system_combo)
        
        # Point group
        self.point_group_edit = QLineEdit()
        self.point_group_edit.setText(self.mineral_data.get('point_group', ''))
        form_layout.addRow("Point Group:", self.point_group_edit)
        
        # Space group
        self.space_group_edit = QLineEdit()
        self.space_group_edit.setText(self.mineral_data.get('space_group', ''))
        form_layout.addRow("Space Group:", self.space_group_edit)
        
        layout.addLayout(form_layout)
        
        # Buttons
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def get_mineral_data(self):
        """Get the mineral data from the dialog."""
        return {
            'name': self.name_edit.text(),
            'crystal_system': self.crystal_system_combo.currentText(),
            'point_group': self.point_group_edit.text(),
            'space_group': self.space_group_edit.text(),
            'modes': self.mineral_data.get('modes', [])
        }


class ModeEditDialog(QDialog):
    """Dialog for editing Raman mode properties."""
    
    def __init__(self, parent=None, mode_data=None):
        super().__init__(parent)
        self.mode_data = mode_data or (0, '', 1.0)
        
        self.setWindowTitle("Edit Mode" if mode_data else "Add Mode")
        self.setMinimumSize(300, 200)
        
        self.setup_ui()
    
    def setup_ui(self):
        """Set up the dialog UI."""
        layout = QVBoxLayout(self)
        
        # Form layout for mode properties
        form_layout = QFormLayout()
        
        # Position
        self.position_spin = QDoubleSpinBox()
        self.position_spin.setRange(0, 4000)
        self.position_spin.setSingleStep(1)
        self.position_spin.setValue(float(self.mode_data[0]) if len(self.mode_data) > 0 else 0)
        form_layout.addRow("Position (cm⁻¹):", self.position_spin)
        
        # Symmetry
        self.symmetry_edit = QLineEdit()
        self.symmetry_edit.setText(str(self.mode_data[1]) if len(self.mode_data) > 1 else '')
        form_layout.addRow("Symmetry:", self.symmetry_edit)
        
        # Intensity
        self.intensity_spin = QDoubleSpinBox()
        self.intensity_spin.setRange(0, 100)
        self.intensity_spin.setSingleStep(0.1)
        self.intensity_spin.setValue(float(self.mode_data[2]) if len(self.mode_data) > 2 else 1.0)
        form_layout.addRow("Intensity:", self.intensity_spin)
        
        layout.addLayout(form_layout)
        
        # Buttons
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def get_mode_data(self):
        """Get the mode data from the dialog."""
        return (
            self.position_spin.value(),
            self.symmetry_edit.text(),
            self.intensity_spin.value()
        )


def main():
    """Main function to run the mineral modes browser."""
    app = QApplication(sys.argv)
    
    window = MineralModesDatabaseQt6()
    window.show()
    
    sys.exit(app.exec())


if __name__ == "__main__":
    main() 